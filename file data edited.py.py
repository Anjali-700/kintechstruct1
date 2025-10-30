import numpy as np
import matplotlib.pyplot as plt
from collections import Counter
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.patches as mpatches
import requests
import io
import warnings
from Bio.PDB import PDBParser, PDBList, Selection, NeighborSearch
from Bio.PDB.Polypeptide import PPBuilder, is_aa
from Bio.SVDSuperimposer import SVDSuperimposer
from Bio import BiopythonWarning
import py3Dmol
import tempfile
import webbrowser
from PIL import Image, ImageTk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import shutil
import seaborn as sns
from scipy import stats

# Suppress Biopython warnings
warnings.simplefilter('ignore', BiopythonWarning)
warnings.filterwarnings('ignore')

class ProteinAnalyzer:
    def __init__(self):
        # Amino acid properties databases
        self.amino_acids = set('ACDEFGHIKLMNPQRSTVWY')
        
        # Molecular weights (Da)
        self.aa_weights = {
            'A': 89.09, 'R': 174.20, 'N': 132.12, 'D': 133.10,
            'C': 121.16, 'E': 147.13, 'Q': 146.15, 'G': 75.07,
            'H': 155.16, 'I': 131.17, 'L': 131.17, 'K': 146.19,
            'M': 149.21, 'F': 165.19, 'P': 115.13, 'S': 105.09,
            'T': 119.12, 'W': 204.23, 'Y': 181.19, 'V': 117.15
        }
        
        # pKa values for ionizable groups
        self.pka_values = {
            'D': 3.9, 'E': 4.3,  # Acidic
            'H': 6.0, 'C': 8.3, 'Y': 10.1,  # Partial charges
            'K': 10.5, 'R': 12.5  # Basic
        }
        
        # Hydrophobicity scales (Kyte-Doolittle)
        self.hydrophobicity = {
            'A': 1.8, 'R': -4.5, 'N': -3.5, 'D': -3.5,
            'C': 2.5, 'E': -3.5, 'Q': -3.5, 'G': -0.4,
            'H': -3.2, 'I': 4.5, 'L': 3.8, 'K': -3.9,
            'M': 1.9, 'F': 2.8, 'P': -1.6, 'S': -0.8,
            'T': -0.7, 'W': -0.9, 'Y': -1.3, 'V': 4.2
        }
        
        # Secondary structure propensity
        self.helix_propensity = {
            'A': 1.45, 'R': 0.79, 'N': 0.73, 'D': 0.98,
            'C': 0.77, 'E': 1.53, 'Q': 1.17, 'G': 0.53,
            'H': 1.24, 'I': 1.00, 'L': 1.34, 'K': 1.07,
            'M': 1.20, 'F': 1.12, 'P': 0.59, 'S': 0.79,
            'T': 0.82, 'W': 1.14, 'Y': 0.61, 'V': 1.14
        }
        
        self.sheet_propensity = {
            'A': 0.97, 'R': 0.90, 'N': 0.65, 'D': 0.80,
            'C': 1.30, 'E': 0.26, 'Q': 1.23, 'G': 0.81,
            'H': 0.71, 'I': 1.60, 'L': 1.22, 'K': 0.74,
            'M': 1.67, 'F': 1.28, 'P': 0.62, 'S': 0.72,
            'T': 1.20, 'W': 1.19, 'Y': 1.29, 'V': 1.65
        }
        
        # Protein half-life prediction based on N-terminal residues (in seconds)
        # Based on the N-end rule pathway
        self.half_life_nt = {
            'M': 72000,    # Methionine - stable
            'S': 10800,    # Serine - ~3 hours
            'A': 10800,    # Alanine - ~3 hours
            'T': 10800,    # Threonine - ~3 hours
            'V': 10800,    # Valine - ~3 hours
            'G': 10800,    # Glycine - ~3 hours
            'C': 7200,     # Cysteine - ~2 hours
            'P': 7200,     # Proline - ~2 hours
            'D': 1800,     # Aspartic acid - ~30 minutes
            'E': 1800,     # Glutamic acid - ~30 minutes
            'N': 1800,     # Asparagine - ~30 minutes
            'Q': 1800,     # Glutamine - ~30 minutes
            'H': 600,      # Histidine - ~10 minutes
            'K': 600,      # Lysine - ~10 minutes
            'R': 600,      # Arginine - ~10 minutes
            'F': 300,      # Phenylalanine - ~5 minutes
            'Y': 300,      # Tyrosine - ~5 minutes
            'W': 300,      # Tryptophan - ~5 minutes
            'L': 300,      # Leucine - ~5 minutes
            'I': 300       # Isoleucine - ~5 minutes
        }
        
        # Torsion angle preferences for Ramachandran plot
        self.ramachandran_regions = {
            'Core alpha': {'phi': (-180, -30), 'psi': (-90, 30)},
            'Core beta': {'phi': (-180, -30), 'psi': (60, 180)},
            'Core left-handed alpha': {'phi': (30, 90), 'psi': (-90, 30)},
            'Allowed': {'phi': (-180, 180), 'psi': (-180, 180)}  # Will be refined
        }

    def parse_fasta(self, fasta_text):
        """Parse FASTA format text and extract protein sequence"""
        lines = fasta_text.strip().split('\n')
        sequence = ""
        description = ""
        
        for line in lines:
            if line.startswith('>'):
                description = line[1:].strip()
            else:
                # Remove any spaces or numbers from sequence
                sequence += re.sub(r'[^A-Za-z]', '', line.upper())
        
        return description, sequence

    def validate_sequence(self, sequence):
        """Check if the sequence contains valid amino acids"""
        invalid_chars = set(sequence) - self.amino_acids
        return len(invalid_chars) == 0, invalid_chars

    def calculate_molecular_weight(self, sequence):
        """Calculate molecular weight in Daltons"""
        total_weight = sum(self.aa_weights[aa] for aa in sequence)
        # Subtract water molecular weight for each peptide bond
        return total_weight - (len(sequence) - 1) * 18.02

    def calculate_isoelectric_point(self, sequence):
        """
        Calculate the isoelectric point using an iterative algorithm.
        This method finds the pH at which the net charge of the protein is zero.
        """
        # More accurate pKa values including termini
        pka_map = {
            'C': 8.18, 'D': 3.65, 'E': 4.25,
            'H': 6.00, 'K': 10.53, 'R': 12.48, 'Y': 10.07
        }
        n_term_pka = 9.69  # N-terminus
        c_term_pka = 2.34  # C-terminus
        
        # Start with a search range for pH
        min_ph, max_ph = 0.0, 14.0
        
        for _ in range(100):  # Iterate to find the pI, 100 iterations is plenty for high precision
            mid_ph = (min_ph + max_ph) / 2
            
            # Calculate net charge at mid_ph
            net_charge = 0.0
            
            # N-terminus charge (positive)
            net_charge += 1 / (1 + 10**(mid_ph - n_term_pka))
            # C-terminus charge (negative)
            net_charge -= 1 / (1 + 10**(c_term_pka - mid_ph))
            
            # Charges from amino acid side chains
            for aa, pka in pka_map.items():
                count = sequence.count(aa)
                if aa in 'KRH':  # Basic residues (positive charge)
                    net_charge += count / (1 + 10**(mid_ph - pka))
                elif aa in 'DECY':  # Acidic residues (negative charge)
                    net_charge -= count / (1 + 10**(pka - mid_ph))

            # Adjust the search range
            if net_charge > 0.001:
                min_ph = mid_ph
            elif net_charge < -0.001:
                max_ph = mid_ph
            else:
                return mid_ph # Found the pI
                
        return mid_ph
    
    def calculate_charge_at_ph(self, sequence, ph=7.0):
        """Calculate net charge at specific pH"""
        charge = 0.0
        
        # N-terminal and C-terminal charges
        charge += 1.0 / (1.0 + 10**(ph - 8.0))  # N-terminal
        charge -= 1.0 / (1.0 + 10**(3.9 - ph))   # C-terminal
        
        # Acidic residues
        charge -= sequence.count('D') / (1.0 + 10**(self.pka_values['D'] - ph))
        charge -= sequence.count('E') / (1.0 + 10**(self.pka_values['E'] - ph))
        
        # Basic residues
        charge += sequence.count('R') / (1.0 + 10**(ph - self.pka_values['R']))
        charge += sequence.count('K') / (1.0 + 10**(ph - self.pka_values['K']))
        charge += sequence.count('H') / (1.0 + 10**(ph - self.pka_values['H']))
        
        return charge

    def calculate_hydrophobicity(self, sequence):
        """Calculate average hydrophobicity"""
        return sum(self.hydrophobicity[aa] for aa in sequence) / len(sequence)

    def calculate_amino_acid_composition(self, sequence):
        """Calculate percentage of each amino acid"""
        counter = Counter(sequence)
        total = len(sequence)
        return {aa: (count / total) * 100 for aa, count in counter.items()}

    def calculate_extinction_coefficient(self, sequence):
        """Calculate molar extinction coefficient"""
        # Extinction coefficients for Trp, Tyr, and Cystine (disulfide bonds)
        # Values for reduced form (no disulfide bonds)
        ext_trp = 5500  # M⁻¹cm⁻¹
        ext_tyr = 1490  # M⁻¹cm⁻¹
        ext_cys = 125   # M⁻¹cm⁻¹ (for cystine)
        
        trp_count = sequence.count('W')
        tyr_count = sequence.count('Y')
        cys_count = sequence.count('C')
        
        return trp_count * ext_trp + tyr_count * ext_tyr + cys_count * ext_cys

    def calculate_instability_index(self, sequence):
        """
        Calculate the instability index using the method of Guruprasad et al., 1990.
        Any value above 40 indicates the protein may be unstable.
        """
        # Dipeptide Instability Weight Value (DIWV) matrix from Guruprasad et al.
        diwv = {
            'A': {'A': 43.33, 'L': 36.67, 'R': 5.56, 'N': -23.33, 'D': -35.56, 'C': 12.22, 'Q': -20.00, 'E': -14.44, 'G': -3.33, 'H': -22.22, 'I': 94.44, 'K': -5.56, 'M': 42.22, 'F': -15.56, 'P': -1.11, 'S': 22.22, 'T': 53.33, 'W': -41.11, 'Y': -22.22, 'V': 80.00},
            'L': {'A': 36.67, 'L': 36.67, 'R': -12.22, 'N': -35.56, 'D': -47.78, 'C': -5.56, 'Q': -32.22, 'E': -26.67, 'G': -15.56, 'H': -34.44, 'I': 94.44, 'K': -17.78, 'M': 30.00, 'F': -27.78, 'P': -13.33, 'S': 10.00, 'T': 41.11, 'W': -53.33, 'Y': -34.44, 'V': 80.00},
            'R': {'A': 5.56, 'L': -12.22, 'R': 5.56, 'N': -17.78, 'D': -30.00, 'C': -1.11, 'Q': -5.56, 'E': 2.22, 'G': -10.00, 'H': -3.33, 'I': -3.33, 'K': 12.22, 'M': -2.22, 'F': -10.00, 'P': 16.67, 'S': 5.56, 'T': 7.78, 'W': 14.44, 'Y': 1.11, 'V': 1.11},
            'N': {'A': -23.33, 'L': -35.56, 'R': -17.78, 'N': -23.33, 'D': -12.22, 'C': -14.44, 'Q': -14.44, 'E': -10.00, 'G': 14.44, 'H': 2.22, 'I': -32.22, 'K': -12.22, 'M': -31.11, 'F': -31.11, 'P': -15.56, 'S': 2.22, 'T': -10.00, 'W': -12.22, 'Y': -2.22, 'V': -31.11},
            'D': {'A': -35.56, 'L': -47.78, 'R': -30.00, 'N': -12.22, 'D': -35.56, 'C': -26.67, 'Q': -26.67, 'E': -3.33, 'G': 1.11, 'H': -11.11, 'I': -44.44, 'K': -24.44, 'M': -43.33, 'F': -43.33, 'P': -27.78, 'S': -10.00, 'T': -22.22, 'W': -24.44, 'Y': -14.44, 'V': -43.33},
            'C': {'A': 12.22, 'L': -5.56, 'R': -1.11, 'N': -14.44, 'D': -26.67, 'C': 12.22, 'Q': -16.67, 'E': -21.11, 'G': -10.00, 'H': -18.89, 'I': 58.89, 'K': -12.22, 'M': -1.11, 'F': 20.00, 'P': -12.22, 'S': 17.78, 'T': 16.67, 'W': -35.56, 'Y': 2.22, 'V': 52.22},
            'Q': {'A': -20.00, 'L': -32.22, 'R': -5.56, 'N': -14.44, 'D': -26.67, 'C': -16.67, 'Q': -5.56, 'E': 4.44, 'G': -10.00, 'H': 4.44, 'I': -28.89, 'K': 1.11, 'M': -25.56, 'F': -25.56, 'P': -1.11, 'S': -6.67, 'T': -18.89, 'W': 2.22, 'Y': -10.00, 'V': -27.78},
            'E': {'A': -14.44, 'L': -26.67, 'R': 2.22, 'N': -10.00, 'D': -3.33, 'C': -21.11, 'Q': 4.44, 'E': -14.44, 'G': -4.44, 'H': -5.56, 'I': -23.33, 'K': 10.00, 'M': -20.00, 'F': -20.00, 'P': 3.33, 'S': -1.11, 'T': -13.33, 'W': 8.89, 'Y': -4.44, 'V': -22.22},
            'G': {'A': -3.33, 'L': -15.56, 'R': -10.00, 'N': 14.44, 'D': 1.11, 'C': -10.00, 'Q': -10.00, 'E': -4.44, 'G': -3.33, 'H': -11.11, 'I': -14.44, 'K': -4.44, 'M': -13.33, 'F': -13.33, 'P': -1.11, 'S': 22.22, 'T': 2.22, 'W': -27.78, 'Y': -10.00, 'V': -13.33},
            'H': {'A': -22.22, 'L': -34.44, 'R': -3.33, 'N': 2.22, 'D': -11.11, 'C': -18.89, 'Q': 4.44, 'E': -5.56, 'G': -11.11, 'H': -3.33, 'I': -31.11, 'K': -2.22, 'M': -30.00, 'F': -11.11, 'P': -15.56, 'S': -6.67, 'T': -18.89, 'W': 11.11, 'Y': 16.67, 'V': -30.00},
            'I': {'A': 94.44, 'L': 94.44, 'R': -3.33, 'N': -32.22, 'D': -44.44, 'C': 58.89, 'Q': -28.89, 'E': -23.33, 'G': -14.44, 'H': -31.11, 'I': 94.44, 'K': -14.44, 'M': 93.33, 'F': 2.22, 'P': -12.22, 'S': 13.33, 'T': 54.44, 'W': -50.00, 'Y': -21.11, 'V': 93.33},
            'K': {'A': -5.56, 'L': -17.78, 'R': 12.22, 'N': -12.22, 'D': -24.44, 'C': -12.22, 'Q': 1.11, 'E': 10.00, 'G': -4.44, 'H': -2.22, 'I': -14.44, 'K': -5.56, 'M': -12.22, 'F': -12.22, 'P': 10.00, 'S': 1.11, 'T': -3.33, 'W': 10.00, 'Y': -3.33, 'V': -13.33},
            'M': {'A': 42.22, 'L': 30.00, 'R': -2.22, 'N': -31.11, 'D': -43.33, 'C': -1.11, 'Q': -25.56, 'E': -20.00, 'G': -13.33, 'H': -30.00, 'I': 93.33, 'K': -12.22, 'M': 42.22, 'F': -13.33, 'P': -13.33, 'S': 10.00, 'T': 41.11, 'W': -50.00, 'Y': -30.00, 'V': 81.11},
            'F': {'A': -15.56, 'L': -27.78, 'R': -10.00, 'N': -31.11, 'D': -43.33, 'C': 20.00, 'Q': -25.56, 'E': -20.00, 'G': -13.33, 'H': -11.11, 'I': 2.22, 'K': -12.22, 'M': -13.33, 'F': -15.56, 'P': -26.67, 'S': -2.22, 'T': -3.33, 'W': 41.11, 'Y': 12.22, 'V': 2.22},
            'P': {'A': -1.11, 'L': -13.33, 'R': 16.67, 'N': -15.56, 'D': -27.78, 'C': -12.22, 'Q': -1.11, 'E': 3.33, 'G': -1.11, 'H': -15.56, 'I': -12.22, 'K': 10.00, 'M': -13.33, 'F': -26.67, 'P': -1.11, 'S': 10.00, 'T': 2.22, 'W': -28.89, 'Y': -15.56, 'V': -11.11},
            'S': {'A': 22.22, 'L': 10.00, 'R': 5.56, 'N': 2.22, 'D': -10.00, 'C': 17.78, 'Q': -6.67, 'E': -1.11, 'G': 22.22, 'H': -6.67, 'I': 13.33, 'K': 1.11, 'M': 10.00, 'F': -2.22, 'P': 10.00, 'S': 22.22, 'T': 28.89, 'W': -20.00, 'Y': 1.11, 'V': 14.44},
            'T': {'A': 53.33, 'L': 41.11, 'R': 7.78, 'N': -10.00, 'D': -22.22, 'C': 16.67, 'Q': -18.89, 'E': -13.33, 'G': 2.22, 'H': -18.89, 'I': 54.44, 'K': -3.33, 'M': 41.11, 'F': -3.33, 'P': 2.22, 'S': 28.89, 'T': 53.33, 'W': -36.67, 'Y': -14.44, 'V': 54.44},
            'W': {'A': -41.11, 'L': -53.33, 'R': 14.44, 'N': -12.22, 'D': -24.44, 'C': -35.56, 'Q': 2.22, 'E': 8.89, 'G': -27.78, 'H': 11.11, 'I': -50.00, 'K': 10.00, 'M': -50.00, 'F': 41.11, 'P': -28.89, 'S': -20.00, 'T': -36.67, 'W': -41.11, 'Y': 44.44, 'V': -48.89},
            'Y': {'A': -22.22, 'L': -34.44, 'R': 1.11, 'N': -2.22, 'D': -14.44, 'C': 2.22, 'Q': -10.00, 'E': -4.44, 'G': -10.00, 'H': 16.67, 'I': -21.11, 'K': -3.33, 'M': -30.00, 'F': 12.22, 'P': -15.56, 'S': 1.11, 'T': -14.44, 'W': 44.44, 'Y': -22.22, 'V': -20.00},
            'V': {'A': 80.00, 'L': 80.00, 'R': 1.11, 'N': -31.11, 'D': -43.33, 'C': 52.22, 'Q': -27.78, 'E': -22.22, 'G': -13.33, 'H': -30.00, 'I': 93.33, 'K': -13.33, 'M': 81.11, 'F': 2.22, 'P': -11.11, 'S': 14.44, 'T': 54.44, 'W': -48.89, 'Y': -20.00, 'V': 80.00},
        }

        if len(sequence) < 2:
            return 0.0
            
        score = 0.0
        for i in range(len(sequence) - 1):
            dipeptide = sequence[i:i+2]
            score += diwv[dipeptide[0]][dipeptide[1]]
        
        return (1/len(sequence)) * score

    def calculate_half_life(self, sequence):
        """
        Predict protein half-life based on the N-terminal residue (N-end rule).
        This implementation uses values derived from S. cerevisiae (yeast).
        Returns the predicted half-life in seconds.
        """
        # N-end rule half-life values for S. cerevisiae (yeast) in seconds
        half_life_map = {
            # Stabilizing residues (>20 hours = 72000s)
            'A': 72000, 'S': 72000, 'T': 72000, 'G': 72000, 'V': 72000, 'C': 72000,
            # Methionine is also stabilizing
            'M': 72000,
            # Proline is complex but generally stabilizing
            'P': 72000,
            # Destabilizing residues
            'I': 1800,  # ~30 min
            'E': 1800,  # ~30 min
            'Y': 600,   # ~10 min
            'Q': 600,   # ~10 min
            'D': 180,   # ~3 min
            'N': 180,   # ~3 min
            'H': 180,   # ~3 min
            'L': 120,   # ~2 min
            'F': 120,   # ~2 min
            'W': 120,   # ~2 min
            'K': 120,   # ~2 min
            'R': 120,   # ~2 min
        }
        
        if not sequence:
            return 0
        
        n_terminal = sequence[0]
        # Default to a stable half-life if the residue is not in the map
        return half_life_map.get(n_terminal, 72000)
    
    def predict_secondary_structure(self, sequence):
        """
        Predict secondary structure using a simple propensity-based method.
        Note: This is a rough estimation. For accurate predictions, use
        specialized tools like PSIPRED or JPred.
        """
        # For very short peptides, stable structures are not possible
        if len(sequence) < 5:
            return 0.0, 0.0, 100.0

        helix_score = 0
        sheet_score = 0
        coil_score = 0

        # Simple heuristic: Proline and Glycine are strong coil/turn formers
        for aa in sequence:
            helix_score += self.helix_propensity[aa]
            sheet_score += self.sheet_propensity[aa]
            if aa in 'GPNDS': # Gly, Pro, Asn, Asp, Ser are common in turns/coils
                coil_score += 1

        # Normalize scores
        total_propensity = helix_score + sheet_score + coil_score
        
        if total_propensity == 0:
            return 33.3, 33.3, 33.4 # Avoid division by zero

        helix_percent = (helix_score / total_propensity) * 100
        sheet_percent = (sheet_score / total_propensity) * 100
        coil_percent = (coil_score / total_propensity) * 100

        # Further normalization to ensure the sum is 100
        total_percent = helix_percent + sheet_percent + coil_percent
        helix = (helix_percent / total_percent) * 100
        sheet = (sheet_percent / total_percent) * 100
        coil = (coil_percent / total_percent) * 100
        
        return helix, sheet, coil

    def generate_ramachandran_plot(self, sequence):
        """Generate phi and psi angles for a Ramachandran plot"""
        # For a real implementation, we would need 3D structure data
        # Since we only have sequence, we'll generate synthetic angles
        # based on amino acid preferences
        
        phi_angles = []
        psi_angles = []
        residue_types = []
        
        # Generate angles based on residue type preferences
        for i, aa in enumerate(sequence):
            if aa in 'ACDEFGHIKLMNPQRSTVWY':
                # Base angle ranges based on residue type
                if aa in 'EAMQK':  # Helix formers
                    phi = np.random.uniform(-70, -40)
                    psi = np.random.uniform(-50, -10)
                elif aa in 'VITYFCW':  # Beta formers
                    phi = np.random.uniform(-150, -100)
                    psi = np.random.uniform(100, 150)
                elif aa == 'G':  # Glycine - very flexible
                    phi = np.random.uniform(-180, 180)
                    psi = np.random.uniform(-180, 180)
                elif aa == 'P':  # Proline - constrained
                    phi = np.random.uniform(-70, -40)
                    psi = np.random.uniform(-180, 180)
                else:  # Others
                    phi = np.random.uniform(-180, 180)
                    psi = np.random.uniform(-180, 180)
                
                # Add some noise
                phi += np.random.normal(0, 10)
                psi += np.random.normal(0, 10)
                
                # Keep angles in [-180, 180] range
                phi = (phi + 180) % 360 - 180
                psi = (psi + 180) % 360 - 180
                
                phi_angles.append(phi)
                psi_angles.append(psi)
                residue_types.append(aa)
        
        return phi_angles, psi_angles, residue_types

    def fetch_pdb_structure(self, protein_id):
        """Fetch PDB structure from RCSB database"""
        try:
            # Create temporary directory for PDB files
            temp_dir = tempfile.mkdtemp()
            pdb_list = PDBList()
            
            # Fetch PDB file
            pdb_file = pdb_list.retrieve_pdb_file(
                protein_id.upper(), 
                file_format='pdb', 
                pdir=temp_dir
            )
            
            if not pdb_file or not os.path.exists(pdb_file):
                raise Exception(f"Could not download PDB file for {protein_id}")
            
            # Parse the structure
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure(protein_id, pdb_file)
            
            # Copy to a more permanent temporary file
            permanent_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdb')
            shutil.copy2(pdb_file, permanent_temp.name)
            
            # Clean up temporary directory
            shutil.rmtree(temp_dir)
            
            return structure, permanent_temp.name
            
        except Exception as e:
            print(f"Error fetching PDB structure: {e}")
            # Clean up on error
            if 'temp_dir' in locals():
                try:
                    shutil.rmtree(temp_dir)
                except:
                    pass
            return None, None

    def fetch_pdb_sequence(self, protein_id):
        """Extract sequence from PDB structure"""
        try:
            structure, pdb_file = self.fetch_pdb_structure(protein_id)
            if not structure:
                return None, None
            
            # Extract sequence using PPBuilder
            ppb = PPBuilder()
            sequences = []
            
            for pp in ppb.build_peptides(structure):
                sequences.append(pp.get_sequence())
            
            if sequences:
                full_sequence = "".join(str(seq) for seq in sequences)
                return full_sequence, pdb_file
            else:
                return None, pdb_file
                
        except Exception as e:
            print(f"Error extracting sequence from PDB: {e}")
            return None, None

    def analyze_pdb_structure(self, pdb_file):
        """Analyze 3D structure from PDB file"""
        try:
            parser = PDBParser(QUIET=True)
            structure = parser.get_structure('analyzed', pdb_file)
            
            analysis = {}
            
            # Get basic structure info
            analysis['num_models'] = len(structure)
            analysis['num_chains'] = 0
            analysis['num_residues'] = 0
            analysis['num_atoms'] = 0
            
            for model in structure:
                analysis['num_chains'] += len(model)
                for chain in model:
                    for residue in chain:
                        if is_aa(residue):
                            analysis['num_residues'] += 1
                            analysis['num_atoms'] += len(residue)
            
            # Calculate center of mass
            atoms = list(structure.get_atoms())
            if atoms:
                coords = np.array([atom.get_coord() for atom in atoms])
                analysis['center_of_mass'] = coords.mean(axis=0)
                analysis['dimensions'] = coords.ptp(axis=0)  # Peak-to-peak (max-min)
            
            return analysis
            
        except Exception as e:
            print(f"Error analyzing PDB structure: {e}")
            return None

    def visualize_structure(self, pdb_file, protein_id):
        """Create 3D structure visualization using py3Dmol"""
        try:
            # Read PDB file
            with open(pdb_file, 'r') as f:
                pdb_data = f.read()
            
            # Create a complete HTML page with embedded 3Dmol.js
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <title>{protein_id} 3D Structure</title>
                <script src="https://cdnjs.cloudflare.com/ajax/libs/3Dmol/2.1.0/3Dmol-min.js"></script>
                <style>
                    body {{ margin: 0; padding: 0; }}
                    #viewer {{ width: 100%; height: 100vh; }}
                    .controls {{ 
                        position: absolute; 
                        top: 10px; 
                        left: 10px; 
                        background: white; 
                        padding: 10px; 
                        border-radius: 5px;
                        box-shadow: 0 2px 5px rgba(0,0,0,0.2);
                    }}
                    button {{ margin: 2px; padding: 5px 10px; }}
                </style>
            </head>
            <body>
                <div class="controls">
                    <button onclick="setStyle({{cartoon: {{}}}})">Cartoon</button>
                    <button onclick="setStyle({{stick: {{}}}})">Sticks</button>
                    <button onclick="setStyle({{sphere: {{radius: 1.5}}}})">Spheres</button>
                    <button onclick="viewer.zoomTo()">Zoom To Fit</button>
                </div>
                <div id="viewer"></div>
                <script>
                    let viewer = $3Dmol.createViewer("viewer");
                    viewer.addModel(`{pdb_data}`, "pdb");
                    
                    function setStyle(style) {{
                        viewer.setStyle({{}}, style);
                        viewer.render();
                    }}
                    
                    // Default to cartoon view
                    setStyle({{cartoon: {{color: "spectrum"}}}});
                </script>
            </body>
            </html>
            """
            
            # Save to temporary file
            temp_html = tempfile.NamedTemporaryFile(delete=False, suffix='.html')
            with open(temp_html.name, 'w') as f:
                f.write(html_content)
            
            return temp_html.name
        except Exception as e:
            print(f"Error creating 3D visualization: {e}")
            return None

    def analyze_sequence(self, sequence):
        """Perform complete analysis of protein sequence"""
        if not sequence:
            return {"error": "Empty sequence provided"}
        
        # Validate sequence
        is_valid, invalid_chars = self.validate_sequence(sequence)
        if not is_valid:
            return {"error": f"Invalid amino acids found: {invalid_chars}"}
        
        # Perform all analyses
        analysis = {
            'length': len(sequence),
            'molecular_weight': self.calculate_molecular_weight(sequence),
            'isoelectric_point': self.calculate_isoelectric_point(sequence),
            'charge_at_ph7': self.calculate_charge_at_ph(sequence, 7.0),
            'hydrophobicity': self.calculate_hydrophobicity(sequence),
            'amino_acid_composition': self.calculate_amino_acid_composition(sequence),
            'extinction_coefficient': self.calculate_extinction_coefficient(sequence),
            'instability_index': self.calculate_instability_index(sequence),
            'half_life': self.calculate_half_life(sequence),
        }
        
        # Add secondary structure prediction
        helix, sheet, coil = self.predict_secondary_structure(sequence)
        analysis['secondary_structure'] = {
            'helix': helix,
            'sheet': sheet,
            'coil': coil
        }
        
        # Generate Ramachandran plot data
        phi_angles, psi_angles, residue_types = self.generate_ramachandran_plot(sequence)
        analysis['ramachandran_data'] = {
            'phi': phi_angles,
            'psi': psi_angles,
            'residues': residue_types
        }
        
        return analysis

    def format_analysis_report(self, analysis, description=""):
        """Format analysis results as a readable report"""
        if 'error' in analysis:
            return f"Error: {analysis['error']}"
        
        report = []
        if description:
            report.append(f"PROTEIN ANALYSIS REPORT: {description}")
        else:
            report.append("PROTEIN ANALYSIS REPORT")
        report.append("=" * 50)
        
        report.append(f"Sequence length: {analysis['length']} amino acids")
        report.append(f"Molecular weight: {analysis['molecular_weight']:.2f} Da")
        report.append(f"Isoelectric point (pI): {analysis['isoelectric_point']:.2f}")
        report.append(f"Net charge at pH 7.0: {analysis['charge_at_ph7']:.2f}")
        report.append(f"Average hydrophobicity: {analysis['hydrophobicity']:.2f}")
        report.append(f"Molar extinction coefficient: {analysis['extinction_coefficient']:.0f} M⁻¹cm⁻¹")
        report.append(f"Instability index: {analysis['instability_index']:.1f} (lower is more stable)")
        
        # Add this line for half-life
        half_life_seconds = analysis['half_life']
        half_life_hours = half_life_seconds / 3600
        report.append(f"Predicted half-life (N-end rule): {half_life_hours:.1f} hours")
        
        # Secondary structure
        ss = analysis['secondary_structure']
        report.append(f"Predicted secondary structure:")
        report.append(f"  α-helix: {ss['helix']:.1f}%")
        report.append(f"  β-sheet: {ss['sheet']:.1f}%")
        report.append(f"  Random coil: {ss['coil']:.1f}%")
        
        # Amino acid composition
        report.append("\nAmino acid composition:")
        comp = analysis['amino_acid_composition']
        sorted_comp = sorted(comp.items(), key=lambda x: x[1], reverse=True)
        
        for i in range(0, len(sorted_comp), 4):
            line = "  "
            for aa, perc in sorted_comp[i:i+4]:
                line += f"{aa}: {perc:.1f}%  "
            report.append(line)
        
        return "\n".join(report)
    
    def export_to_excel(self, analysis, description="", filename="protein_analysis.xlsx"):
        """Export analysis results to an Excel file"""
        try:
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Protein Analysis"
            
            # Add header with title and date
            ws['A1'] = "PROTEIN ANALYSIS REPORT"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            if description:
                ws['A3'] = f"Protein: {description}"
                ws['A3'].font = Font(bold=True)
            
            # Add basic properties
            ws['A5'] = "Basic Properties"
            ws['A5'].font = Font(bold=True, size=14)
            
            basic_data = [
                ["Property", "Value"],
                ["Sequence length", f"{analysis['length']} amino acids"],
                ["Molecular weight", f"{analysis['molecular_weight']:.2f} Da"],
                ["Isoelectric point (pI)", f"{analysis['isoelectric_point']:.2f}"],
                ["Net charge at pH 7.0", f"{analysis['charge_at_ph7']:.2f}"],
                ["Average hydrophobicity", f"{analysis['hydrophobicity']:.2f}"],
                ["Molar extinction coefficient", f"{analysis['extinction_coefficient']:.0f} M⁻¹cm⁻¹"],
                ["Instability index", f"{analysis['instability_index']:.1f} (lower is more stable)"],
                ["Predicted half-life", f"{analysis['half_life']/3600:.1f} hours"]
            ]
            
            for row_idx, row_data in enumerate(basic_data, start=6):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 6:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add secondary structure
            ss_start_row = 6 + len(basic_data) + 2
            ws.cell(row=ss_start_row, column=1, value="Secondary Structure Prediction").font = Font(bold=True, size=14)
            
            ss_data = [
                ["Structure Type", "Percentage"],
                ["α-helix", f"{analysis['secondary_structure']['helix']:.1f}%"],
                ["β-sheet", f"{analysis['secondary_structure']['sheet']:.1f}%"],
                ["Random coil", f"{analysis['secondary_structure']['coil']:.1f}%"]
            ]
            
            for row_idx, row_data in enumerate(ss_data, start=ss_start_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == ss_start_row + 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add amino acid composition
            comp_start_row = ss_start_row + len(ss_data) + 3
            ws.cell(row=comp_start_row, column=1, value="Amino Acid Composition").font = Font(bold=True, size=14)
            
            comp_data = [["Amino Acid", "Percentage"]]
            comp = analysis['amino_acid_composition']
            sorted_comp = sorted(comp.items(), key=lambda x: x[1], reverse=True)
            
            for aa, perc in sorted_comp:
                comp_data.append([aa, f"{perc:.1f}%"])
            
            for row_idx, row_data in enumerate(comp_data, start=comp_start_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == comp_start_row + 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add Ramachandran plot data
            rama_start_row = comp_start_row + len(comp_data) + 3
            ws.cell(row=rama_start_row, column=1, value="Ramachandran Plot Data").font = Font(bold=True, size=14)
            
            rama_data = [["Residue", "Phi Angle", "Psi Angle"]]
            for i, (res, phi, psi) in enumerate(zip(analysis['ramachandran_data']['residues'],
                                                  analysis['ramachandran_data']['phi'],
                                                  analysis['ramachandran_data']['psi'])):
                rama_data.append([res, f"{phi:.1f}", f"{psi:.1f}"])
                if i >= 50:  # Limit to first 50 residues for Excel
                    rama_data.append(["...", "...", "..."])
                    break
            
            for row_idx, row_data in enumerate(rama_data, start=rama_start_row + 1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == rama_start_row + 1:  # Header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
            
class EnzymeKineticsAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Enzyme Kinetics Database Analyzer")
        self.root.geometry("1600x1000")
        self.root.configure(bg='#f0f0f0')
        
        # Load the data
        try:
            self.df = pd.read_csv('all_enzyme_kcat_km_ddG_all_clean_pH.csv')
            messagebox.showinfo("Success", f"Dataset loaded successfully!\nTotal entries: {len(self.df):,}")
        except FileNotFoundError:
            messagebox.showerror("File Not Found", "Database file 'all_enzyme_kcat_km_ddG_all_clean_pH.csv' not found.")
            return
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load database: {str(e)}")
            return
        
        # Preprocess the data
        self.preprocess_data()
        
        # Create GUI elements
        self.create_gui()
        
        # Store current selection
        self.current_enzyme_data = None
        
    def create_menu(self):
        """Create menu bar with help option"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
    
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Export Data", command=self.export_data)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
    
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="User Guide", command=self.show_user_guide)
        help_menu.add_command(label="About", command=self.show_about)
        
    def show_user_guide(self):
        """Display comprehensive user guide"""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("User Guide - Enzyme Kinetics Analyzer")
        guide_window.geometry("900x700")
        
        guide_text = scrolledtext.ScrolledText(guide_window, wrap=tk.WORD, width=100, height=50)
        guide_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        guide_content = """

ADVANCED ENZYME KINETICS ANALYZER - USER GUIDE
==============================================

1. QUICK SEARCH
---------------
- Use the unified search box at the top to search by EC number, substrate, or organism
- Select search type from dropdown and enter your search term
- Results will show in the Search Results tab

2. SEARCH RESULTS TAB
---------------------
- View filtered results from your search
- Click on any entry to view detailed enzyme information
- Columns show key kinetic parameters and changes

3. DATA VISUALIZATION TAB
-------------------------
- Generate various plots to analyze enzyme kinetics data:
  * kcat Comparison: Bar chart comparing mutant vs wild-type kcat values
  * Km Comparison: Scatter plot of mutant vs wild-type Km (log scale)
  * ΔΔG Distribution: Histogram of stability changes
  * Temperature Effect: Relationship between temperature and kcat
  * Substrate Specificity: Top substrates by specificity constant
  * Efficiency Correlation: Relationship between ΔΔG and catalytic efficiency
  * Mutation Impact: Effect of different mutation types on kinetics
  * Enzyme Class Comparison: Compare kinetics across EC classes
  * pH Effect: Relationship between pH and enzyme activity
  * kcat/Km Ratio: Distribution of catalytic efficiency

4. ENZYME INFORMATION TAB
-------------------------
- Displays comprehensive details for selected enzymes including:
  * Basic information (EC number, organism, substrate, mutation)
  * Kinetic parameters (kcat, Km for both mutant and wild-type)
  * Thermodynamic parameters (ΔΔG, ΔG values)
  * Calculated parameters (ratios, specificity constants)
  * Experimental conditions and database information

5. STATISTICAL ANALYSIS TAB
---------------------------
- Provides advanced statistical analysis including:
  * Descriptive statistics for all kinetic parameters
  * Correlation analysis between variables
  * Regression analysis and significance testing
  * Normality tests and distribution analysis

6. ADVANCED ANALYSIS TAB
------------------------
- Advanced data analysis tools:
  * Correlation Matrix: Visualize relationships between all parameters
  * Mutation Impact Analysis: Detailed analysis of mutation effects
  * Enzyme Class Comparison: Statistical comparison across EC classes
  * Outlier Detection: Identify statistical outliers in the data

7. DATABASE EXPLORER TAB
------------------------
- Browse the complete dataset
- Load all available entries to explore the database
- Click on any entry to view detailed information
- Export the complete dataset for external analysis

8. ADD NEW DATA TAB
-------------------
- Add new enzyme kinetics data to the database
- Input mutant and wild-type parameters
- Calculate derived parameters automatically
- Validate data before adding to the dataset

KEY PARAMETERS EXPLAINED:
-------------------------
- kcat: Turnover number (reactions per second per enzyme molecule)
- Km: Michaelis constant (substrate concentration at half Vmax)
- ΔΔG: Change in folding free energy due to mutation (kcal/mol)
- Specificity Constant: kcat/Km (measure of catalytic efficiency)
- kcat Ratio: Mutant kcat / Wild-type kcat
- Km Ratio: Mutant Km / Wild-type Km

TIPS FOR EFFECTIVE USE:
-----------------------
1. Start with the Database Explorer to understand available data
2. Use Quick Search for specific enzyme queries
3. Click on results to view detailed enzyme information
4. Use multiple plot types to understand data patterns
5. Check statistical significance in the Statistics tab
6. Use Advanced Analysis for comprehensive data exploration

For scientific publications, always validate findings with appropriate statistical tests.
"""
        guide_text.insert(1.0, guide_content)
        guide_text.config(state=tk.DISABLED)
        
    def show_about(self):
        """Display about information"""
        about_text = """
Advanced Enzyme Kinetics Analyzer

Version: 1.0
Developed for: Enzyme Kinetics Research and Protein Engineering

Features:
- Comprehensive database exploration
- Advanced search and filtering
- Multiple visualization types
- Detailed statistical analysis
- Mutation impact assessment
- Export capabilities

This tool is designed for researchers studying enzyme kinetics, 
protein engineering, and mutational effects on enzyme function.
"""
        messagebox.showinfo("About", about_text)
        
    def preprocess_data(self):
        """Preprocess the data for better analysis"""
        # Convert numeric columns, handling errors
        numeric_columns = ['kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG', 'temperature', 'pH']
        for col in numeric_columns:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors='coerce')
        
        # Create derived columns with error handling
        if all(col in self.df.columns for col in ['kcat_mut', 'kcat_wt']):
            self.df['kcat_ratio'] = self.df['kcat_mut'] / self.df['kcat_wt']
            # Handle infinite values
            self.df['kcat_ratio'] = self.df['kcat_ratio'].replace([np.inf, -np.inf], np.nan)
        
        if all(col in self.df.columns for col in ['Km_mut', 'Km_wt']):
            self.df['Km_ratio'] = self.df['Km_mut'] / self.df['Km_wt']
            self.df['Km_ratio'] = self.df['Km_ratio'].replace([np.inf, -np.inf], np.nan)
        
        if all(col in self.df.columns for col in ['kcat_mut', 'Km_mut']):
            self.df['specificity_constant_mut'] = self.df['kcat_mut'] / self.df['Km_mut']
            self.df['specificity_constant_mut'] = self.df['specificity_constant_mut'].replace([np.inf, -np.inf], np.nan)
        
        if all(col in self.df.columns for col in ['kcat_wt', 'Km_wt']):
            self.df['specificity_constant_wt'] = self.df['kcat_wt'] / self.df['Km_wt']
            self.df['specificity_constant_wt'] = self.df['specificity_constant_wt'].replace([np.inf, -np.inf], np.nan)
        
        if all(col in self.df.columns for col in ['specificity_constant_mut', 'specificity_constant_wt']):
            self.df['catalytic_efficiency_change'] = self.df['specificity_constant_mut'] / self.df['specificity_constant_wt']
            self.df['catalytic_efficiency_change'] = self.df['catalytic_efficiency_change'].replace([np.inf, -np.inf], np.nan)
        
        # Calculate log values for better distribution
        if 'kcat_ratio' in self.df.columns:
            self.df['log_kcat_ratio'] = np.log10(self.df['kcat_ratio'].clip(lower=1e-10))
        
        if 'Km_ratio' in self.df.columns:
            self.df['log_Km_ratio'] = np.log10(self.df['Km_ratio'].clip(lower=1e-10))
        
        # Fill NaN values in critical columns
        if 'organism' in self.df.columns:
            self.df['organism'] = self.df['organism'].fillna('Unknown')
        if 'substrate_kinetics' in self.df.columns:
            self.df['substrate_kinetics'] = self.df['substrate_kinetics'].fillna('Unknown substrate')
        
        # Create mutation type categories if possible
        if 'wt_aa_type' in self.df.columns and 'mut_aa_type' in self.df.columns:
            self.df['mutation_category'] = self.df.apply(
                lambda x: f"{x['wt_aa_type']}→{x['mut_aa_type']}", axis=1
            )
        
        # Create simplified EC class (first number)
        if 'ec_number' in self.df.columns:
            self.df['ec_class'] = self.df['ec_number'].str.split('.').str[0]
        
    def create_gui(self):
        """Create the main GUI interface"""
        # Create simplified search box at the top
        self.create_unified_search()
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create tabs in desired order - Add New Data is now last
        self.search_tab = ttk.Frame(self.notebook)
        self.visualization_tab = ttk.Frame(self.notebook)
        self.enzyme_info_tab = ttk.Frame(self.notebook)
        self.stats_tab = ttk.Frame(self.notebook)
        self.advanced_tab = ttk.Frame(self.notebook)
        self.database_tab = ttk.Frame(self.notebook)
        self.data_entry_tab = ttk.Frame(self.notebook)  # Now last tab
        
        self.notebook.add(self.search_tab, text="🔍 Search Results")
        self.notebook.add(self.visualization_tab, text="📊 Data Visualization")
        self.notebook.add(self.enzyme_info_tab, text="🧬 Enzyme Information")
        self.notebook.add(self.stats_tab, text="📈 Statistical Analysis")
        self.notebook.add(self.advanced_tab, text="⚡ Advanced Analysis")
        self.notebook.add(self.database_tab, text="🗃️ Database Explorer")
        self.notebook.add(self.data_entry_tab, text="➕ Add New Data")  # Now last
        
        # Setup each tab
        self.setup_search_tab()
        self.setup_visualization_tab()
        self.setup_enzyme_info_tab()
        self.setup_stats_tab()
        self.setup_advanced_tab()
        self.setup_database_tab()
        self.setup_data_entry_tab()  # Setup the data entry tab
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set(f"Ready | Total entries: {len(self.df):,} | Unique enzymes: {self.df['ec_number'].nunique() if 'ec_number' in self.df.columns else 'N/A'}")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief='sunken', anchor='w')
        status_bar.pack(side='bottom', fill='x')
        
    def create_unified_search(self):
        """Create a unified search box at the top of the window"""
        search_frame = ttk.LabelFrame(self.root, text="Quick Search", padding=10)
        search_frame.pack(fill='x', padx=10, pady=5)
        
        # Search type
        ttk.Label(search_frame, text="Search By:").grid(row=0, column=0, sticky='w', padx=5)
        self.search_type = tk.StringVar(value="EC Number")
        search_type_combo = ttk.Combobox(search_frame, textvariable=self.search_type, 
                                       values=["EC Number", "Substrate", "Organism", "Mutation"],
                                       width=12, state="readonly")
        search_type_combo.grid(row=0, column=1, padx=5)
        search_type_combo.bind('<<ComboboxSelected>>', self.update_search_options)
        
        # Search term
        ttk.Label(search_frame, text="Search Term:").grid(row=0, column=2, sticky='w', padx=5)
        self.search_term = tk.StringVar()
        self.search_combo = ttk.Combobox(search_frame, textvariable=self.search_term, width=40)
        self.search_combo.grid(row=0, column=3, padx=5)
        self.search_combo.bind('<Return>', lambda e: self.unified_search())
        
        # Search button
        ttk.Button(search_frame, text="🔍 Search", 
                  command=self.unified_search).grid(row=0, column=4, padx=10)
        
        # Clear button
        ttk.Button(search_frame, text="🗑️ Clear", 
                  command=self.clear_search).grid(row=0, column=5, padx=5)
        
        # Initialize search options
        self.update_search_options()
        
    def update_search_options(self, event=None):
        """Update the search combobox based on selected search type"""
        search_type = self.search_type.get()
        
        try:
            if search_type == "EC Number" and 'ec_number' in self.df.columns:
                options = sorted(self.df['ec_number'].dropna().unique())
            elif search_type == "Substrate" and 'substrate_kinetics' in self.df.columns:
                options = sorted(self.df['substrate_kinetics'].dropna().unique())
            elif search_type == "Organism" and 'organism' in self.df.columns:
                options = sorted(self.df['organism'].dropna().unique())
            elif search_type == "Mutation" and 'clean_mut_wt' in self.df.columns:
                options = sorted(self.df['clean_mut_wt'].dropna().unique())
            else:
                options = []
                
            self.search_combo['values'] = options
            self.search_term.set("")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update search options: {str(e)}")
        
    def unified_search(self):
        """Perform search using the unified search box"""
        search_type = self.search_type.get()
        search_term = self.search_term.get().strip()
        
        if not search_term:
            messagebox.showwarning("Input Required", "Please enter a search term")
            return
        
        try:
            # Filter data based on search type
            filtered_df = self.df.copy()
            
            if search_type == "EC Number" and 'ec_number' in self.df.columns:
                filtered_df = filtered_df[filtered_df['ec_number'] == search_term]
            elif search_type == "Substrate" and 'substrate_kinetics' in self.df.columns:
                filtered_df = filtered_df[filtered_df['substrate_kinetics'].str.contains(search_term, case=False, na=False)]
            elif search_type == "Organism" and 'organism' in self.df.columns:
                filtered_df = filtered_df[filtered_df['organism'].str.contains(search_term, case=False, na=False)]
            elif search_type == "Mutation" and 'clean_mut_wt' in self.df.columns:
                filtered_df = filtered_df[filtered_df['clean_mut_wt'].str.contains(search_term, case=False, na=False)]
            
            if filtered_df.empty:
                messagebox.showinfo("No Results", f"No enzymes found matching {search_type}: {search_term}")
                return
            
            # Switch to search tab and display results
            self.notebook.select(0)
            
            # Clear previous results
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Populate treeview with results
            for _, row in filtered_df.iterrows():
                self.tree.insert('', 'end', values=(
                    row.get('ec_number', 'N/A'),
                    row.get('organism', 'N/A'),
                    str(row.get('substrate_kinetics', 'N/A'))[:30] + '...' if len(str(row.get('substrate_kinetics', 'N/A'))) > 30 else row.get('substrate_kinetics', 'N/A'),
                    row.get('clean_mut_wt', 'N/A'),
                    f"{row.get('kcat_ratio', 'N/A'):.4f}" if pd.notna(row.get('kcat_ratio')) else 'N/A',
                    f"{row.get('Km_ratio', 'N/A'):.4f}" if pd.notna(row.get('Km_ratio')) else 'N/A',
                    f"{row.get('ddG', 'N/A'):.4f}" if pd.notna(row.get('ddG')) else 'N/A',
                    f"{row.get('catalytic_efficiency_change', 'N/A'):.4f}" if pd.notna(row.get('catalytic_efficiency_change')) else 'N/A'
                ))
            
            self.status_var.set(f"Found {len(filtered_df)} matching enzymes | Search: {search_type}: {search_term}")
            messagebox.showinfo("Search Complete", f"Found {len(filtered_df)} matching enzymes")
            
        except Exception as e:
            messagebox.showerror("Search Error", f"Failed to perform search: {str(e)}")
        
    def clear_search(self):
        """Clear the search fields"""
        self.search_term.set("")
        self.status_var.set("Search cleared")

        
    def setup_search_tab(self):
        """Setup the search results tab"""
        # Results frame
        results_frame = ttk.LabelFrame(self.search_tab, text="Search Results", padding=10)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Treeview for results
        columns = ('ec_number', 'organism', 'substrate', 'mutation', 'kcat_ratio', 'Km_ratio', 'ddG', 'efficiency_change')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=20)
        
        # Define headings
        headings = {
            'ec_number': 'EC Number',
            'organism': 'Organism', 
            'substrate': 'Substrate',
            'mutation': 'Mutation',
            'kcat_ratio': 'kcat Ratio',
            'Km_ratio': 'Km Ratio', 
            'ddG': 'ΔΔG',
            'efficiency_change': 'Efficiency Change'
        }
        
        for col, text in headings.items():
            self.tree.heading(col, text=text)
        
        # Configure column widths
        self.tree.column('ec_number', width=120)
        self.tree.column('organism', width=180)
        self.tree.column('substrate', width=250)
        self.tree.column('mutation', width=120)
        self.tree.column('kcat_ratio', width=100)
        self.tree.column('Km_ratio', width=100)
        self.tree.column('ddG', width=100)
        self.tree.column('efficiency_change', width=120)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(results_frame, orient='vertical', command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(results_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack treeview and scrollbars
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        # Bind selection event
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        self.tree.bind('<Double-1>', self.on_tree_double_click)

    def on_tree_select(self, event):
        """Handle treeview selection event"""
        selection = self.tree.selection()
        if selection:
            item = selection[0]
            values = self.tree.item(item, 'values')
            self.display_selected_enzyme(values)

    def on_tree_double_click(self, event):
        """Handle treeview double click - switch to enzyme info tab"""
        selection = self.tree.selection()
        if selection:
            self.notebook.select(2)  # Switch to Enzyme Information tab

    def display_selected_enzyme(self, values):
        """Display enzyme information based on treeview selection"""
        if not values:
            return
            
        try:
            # Find the corresponding row in dataframe
            ec_number = values[0]
            organism = values[1]
            substrate_fragment = values[2].replace('...', '') if '...' in values[2] else values[2]
            mutation = values[3]
            
            # --- FIX: More robust mask ---
            # Start with basic mask
            mask = (self.df['ec_number'] == ec_number) & (self.df['organism'] == organism)
            
            # Handle substrate
            if substrate_fragment == 'N/A':
                # If tree shows N/A, match rows where substrate is null/NaN
                mask = mask & (self.df['substrate_kinetics'].isna())
            else:
                # Otherwise, use 'contains' for the fragment
                mask = mask & (self.df['substrate_kinetics'].str.contains(substrate_fragment, na=False))
            
            # Handle mutation
            if mutation == 'N/A':
                # If tree shows N/A, match rows where mutation is null/NaN
                mask = mask & (self.df['clean_mut_wt'].isna())
            else:
                # Otherwise, exact match
                mask = mask & (self.df['clean_mut_wt'] == mutation)
            # --- END FIX ---
            
            if mask.any():
                row = self.df[mask].iloc[0] # Get the first match
                self.current_enzyme_data = row
                self.display_detailed_enzyme_info(row)
                self.status_var.set(f"Displaying: {ec_number} | {organism} | {substrate_fragment[:30]}...")
            else:
                # --- ADDED: Handle no match ---
                self.status_var.set(f"Could not find details for: {ec_number} | {mutation}")
                
        except Exception as e:
            print(f"Error in display_selected_enzyme: {e}") # For debugging
            messagebox.showerror("Display Error", f"Failed to display enzyme details: {str(e)}")

    def setup_visualization_tab(self):
        """Setup the data visualization tab with advanced options"""
        # Control frame
        control_frame = ttk.LabelFrame(self.visualization_tab, text="Plot Controls", padding=10)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Plot type selection
        ttk.Label(control_frame, text="Plot Type:").grid(row=0, column=0, sticky='w', padx=5)
        self.plot_type = tk.StringVar(value="kcat_comparison")
        plot_combo = ttk.Combobox(control_frame, textvariable=self.plot_type,
                                 values=[
                                     "kcat_comparison", "Km_comparison", "ddG_distribution", 
                                     "temperature_effect", "substrate_specificity", "efficiency_correlation",
                                     "mutation_impact", "enzyme_class_comparison", "ph_effect",
                                     "kcat_km_ratio", "correlation_matrix", "kinetic_parameters_distribution"
                                 ],
                                 width=25, state="readonly")
        plot_combo.grid(row=0, column=1, padx=5, pady=2)
        
        # EC number filter for plots
        ttk.Label(control_frame, text="EC Filter:").grid(row=0, column=2, sticky='w', padx=5)
        self.plot_ec_var = tk.StringVar()
        plot_ec_combo = ttk.Combobox(control_frame, textvariable=self.plot_ec_var,
                                    values=["All"] + sorted(self.df['ec_number'].dropna().unique()),
                                    width=15)
        plot_ec_combo.grid(row=0, column=3, padx=5, pady=2)
        plot_ec_combo.set("All")
        
        # Organism filter
        ttk.Label(control_frame, text="Organism:").grid(row=0, column=4, sticky='w', padx=5)
        self.plot_org_var = tk.StringVar()
        plot_org_combo = ttk.Combobox(control_frame, textvariable=self.plot_org_var,
                                     values=["All"] + sorted(self.df['organism'].dropna().unique()),
                                     width=15)
        plot_org_combo.grid(row=0, column=5, padx=5, pady=2)
        plot_org_combo.set("All")
        
        # Plot button
        ttk.Button(control_frame, text="📈 Generate Plot", 
                  command=self.generate_plot).grid(row=0, column=6, padx=10)
        
        # Clear button
        ttk.Button(control_frame, text="🗑️ Clear Plot", 
                  command=self.clear_plot).grid(row=0, column=7, padx=5)
        
        # Plot description
        self.plot_desc_var = tk.StringVar()
        self.plot_desc_var.set("Select a plot type and click Generate Plot")
        ttk.Label(control_frame, textvariable=self.plot_desc_var, 
                 foreground='blue', font=('Arial', 9)).grid(row=1, column=0, columnspan=8, sticky='w', pady=5)
        
        # Plot frame
        self.plot_frame = ttk.Frame(self.visualization_tab)
        self.plot_frame.pack(fill='both', expand=True, padx=10, pady=5)

    def generate_plot(self):
        """Generate plots based on selected type"""
        # Clear previous plot
        for widget in self.plot_frame.winfo_children():
            widget.destroy()
        
        plot_type = self.plot_type.get()
        ec_filter = self.plot_ec_var.get()
        org_filter = self.plot_org_var.get()
        
        # Filter data
        plot_df = self.df.copy()
        if ec_filter != "All":
            plot_df = plot_df[plot_df['ec_number'] == ec_filter]
        if org_filter != "All":
            plot_df = plot_df[plot_df['organism'] == org_filter]
        
        if plot_df.empty:
            messagebox.showwarning("No Data", "No data found for the selected criteria.")
            return
        
        # Update plot description
        descriptions = {
            "kcat_comparison": "Comparison of mutant vs wild-type kcat values",
            "Km_comparison": "Scatter plot of mutant vs wild-type Km values (log scale)",
            "ddG_distribution": "Distribution of stability changes (ΔΔG) due to mutations",
            "temperature_effect": "Relationship between temperature and catalytic activity",
            "substrate_specificity": "Top substrates by specificity constant (kcat/Km)",
            "efficiency_correlation": "Correlation between ΔΔG and catalytic efficiency changes",
            "mutation_impact": "Impact of different mutation types on enzyme kinetics",
            "enzyme_class_comparison": "Comparison of kinetic parameters across enzyme classes",
            "ph_effect": "Effect of pH on enzyme activity and stability",
            "kcat_km_ratio": "Distribution of catalytic efficiency (kcat/Km) ratios",
            "correlation_matrix": "Correlation between all kinetic parameters",
            "kinetic_parameters_distribution": "Distribution of key kinetic parameters"
        }
        self.plot_desc_var.set(descriptions.get(plot_type, "Plot description not available"))
        
        try:
            fig, ax = plt.subplots(figsize=(12, 8))
            
            plot_methods = {
                "kcat_comparison": self.plot_kcat_comparison,
                "Km_comparison": self.plot_km_comparison,
                "ddG_distribution": self.plot_ddg_distribution,
                "temperature_effect": self.plot_temperature_effect,
                "substrate_specificity": self.plot_substrate_specificity,
                "efficiency_correlation": self.plot_efficiency_correlation,
                "mutation_impact": self.plot_mutation_impact,
                "enzyme_class_comparison": self.plot_enzyme_class_comparison,
                "ph_effect": self.plot_ph_effect,
                "kcat_km_ratio": self.plot_kcat_km_ratio,
                "correlation_matrix": self.plot_correlation_matrix,
                "kinetic_parameters_distribution": self.plot_kinetic_parameters_distribution
            }
            
            if plot_type in plot_methods:
                plot_methods[plot_type](plot_df, ax)
            else:
                ax.text(0.5, 0.5, 'Plot type not implemented', 
                       ha='center', va='center', transform=ax.transAxes, fontsize=14)
            
            plt.tight_layout()
            
            # Embed plot in tkinter
            canvas = FigureCanvasTkAgg(fig, self.plot_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            
            # Add toolbar
            from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
            toolbar = NavigationToolbar2Tk(canvas, self.plot_frame)
            toolbar.update()
            canvas.get_tk_widget().pack(fill='both', expand=True)
            
            self.status_var.set(f"Generated {plot_type} plot | Data points: {len(plot_df)}")
            
        except Exception as e:
            messagebox.showerror("Plot Error", f"Failed to generate plot: {str(e)}")
            # Create error message plot
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, f'Error generating plot:\n{str(e)}', 
                   ha='center', va='center', transform=ax.transAxes, fontsize=12, color='red')
            canvas = FigureCanvasTkAgg(fig, self.plot_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

    def clear_plot(self):
        """Clear the current plot"""
        for widget in self.plot_frame.winfo_children():
            widget.destroy()
        self.plot_desc_var.set("Plot cleared")
        self.status_var.set("Plot area cleared")

    # Enhanced Plot Methods
    def plot_kcat_comparison(self, df, ax):
        """Plot kcat comparison between mutant and wild-type"""
        valid_data = df.dropna(subset=['kcat_mut', 'kcat_wt']).head(15)  # Limit for clarity
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid kcat data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        x = range(len(valid_data))
        width = 0.35
        
        # Create bars
        bars1 = ax.bar([i - width/2 for i in x], valid_data['kcat_wt'], width, 
                      label='Wild-type', alpha=0.7, color='skyblue', edgecolor='black')
        bars2 = ax.bar([i + width/2 for i in x], valid_data['kcat_mut'], width, 
                      label='Mutant', alpha=0.7, color='lightcoral', edgecolor='black')
        
        # Add value labels on bars
        for bars in [bars1, bars2]:
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:.1f}', ha='center', va='bottom', fontsize=8)
        
        ax.set_xlabel('Enzyme Cases')
        ax.set_ylabel('kcat value (s⁻¹)')
        ax.set_title('kcat Comparison: Mutant vs Wild-type\n(Bars show actual kcat values)')
        
        # Create informative labels
        labels = []
        for _, row in valid_data.iterrows():
            label = f"{row.get('ec_number', '')}\n{row.get('clean_mut_wt', '')[:10]}"
            labels.append(label)
        
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Add summary statistics
        if len(valid_data) > 0:
            avg_ratio = valid_data['kcat_ratio'].mean() if 'kcat_ratio' in valid_data.columns else 0
            ax.text(0.02, 0.98, f'Average kcat ratio (mut/wt): {avg_ratio:.2f}', 
                   transform=ax.transAxes, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="white"))

    def plot_km_comparison(self, df, ax):
        """Plot Km comparison between mutant and wild-type (log scale)"""
        valid_data = df.dropna(subset=['Km_mut', 'Km_wt'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid Km data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create scatter plot
        scatter = ax.scatter(valid_data['Km_wt'], valid_data['Km_mut'], 
                           alpha=0.6, c=valid_data.get('ddG', 1), cmap='viridis', s=60)
        
        # Add colorbar if ddG data is available
        if 'ddG' in valid_data.columns and not valid_data['ddG'].isna().all():
            plt.colorbar(scatter, ax=ax, label='ΔΔG (kcal/mol)')
        
        # Add identity line
        lims = [
            np.min([ax.get_xlim(), ax.get_ylim()]),
            np.max([ax.get_xlim(), ax.get_ylim()]),
        ]
        ax.plot(lims, lims, 'k-', alpha=0.75, zorder=0, label='Identity line')
        ax.set_xscale('log')
        ax.set_yscale('log')
        
        ax.set_xlabel('Wild-type Km (log scale)')
        ax.set_ylabel('Mutant Km (log scale)')
        ax.set_title('Km Comparison: Mutant vs Wild-type\n(Colored by ΔΔG where available)')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Add statistics
        km_ratio = valid_data['Km_mut'].mean() / valid_data['Km_wt'].mean() if valid_data['Km_wt'].mean() > 0 else 0
        ax.text(0.02, 0.98, f'Average Km ratio (mut/wt): {km_ratio:.2f}', 
               transform=ax.transAxes, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="white"))

    def plot_ddg_distribution(self, df, ax):
        """Plot distribution of ΔΔG values"""
        valid_data = df['ddG'].dropna()
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid ΔΔG data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create histogram
        n, bins, patches = ax.hist(valid_data, bins=30, alpha=0.7, edgecolor='black', density=True)
        
        # Color bars by stability effect
        for i, (patch, bin_val) in enumerate(zip(patches, bins)):
            if bin_val < 0:
                patch.set_facecolor('green')  # Stabilizing
            else:
                patch.set_facecolor('red')   # Destabilizing
        
        # Add distribution curve
        from scipy.stats import gaussian_kde
        try:
            kde = gaussian_kde(valid_data)
            x_range = np.linspace(valid_data.min(), valid_data.max(), 100)
            ax.plot(x_range, kde(x_range), 'b-', linewidth=2, label='Density')
        except:
            pass
        
        ax.axvline(0, color='black', linestyle='--', alpha=0.8, linewidth=2, label='Stability threshold')
        ax.axvline(valid_data.mean(), color='orange', linestyle='-', alpha=0.8, linewidth=2, label=f'Mean: {valid_data.mean():.2f}')
        
        ax.set_xlabel('ΔΔG (kcal/mol)')
        ax.set_ylabel('Density')
        ax.set_title('Distribution of Stability Changes (ΔΔG)\nGreen: Stabilizing, Red: Destabilizing')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Add statistics
        ax.text(0.02, 0.98, f'Mean: {valid_data.mean():.3f} ± {valid_data.std():.3f} kcal/mol\n'
                           f'Stabilizing: {(valid_data < 0).sum()} | Destabilizing: {(valid_data >= 0).sum()}',
               transform=ax.transAxes, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="white"),
               verticalalignment='top')

    def plot_temperature_effect(self, df, ax):
        """Plot relationship between temperature and kcat"""
        valid_data = df.dropna(subset=['temperature', 'kcat_mut'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid temperature-kcat data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create scatter plot
        scatter = ax.scatter(valid_data['temperature'], valid_data['kcat_mut'], 
                           alpha=0.6, s=50, c=valid_data.get('pH', 7), cmap='plasma')
        
        # Add colorbar for pH
        if 'pH' in valid_data.columns and not valid_data['pH'].isna().all():
            plt.colorbar(scatter, ax=ax, label='pH')
        
        # Add trend line
        try:
            z = np.polyfit(valid_data['temperature'], valid_data['kcat_mut'], 1)
            p = np.poly1d(z)
            ax.plot(valid_data['temperature'], p(valid_data['temperature']), "r--", alpha=0.8, label='Trend line')
        except:
            pass
        
        ax.set_xlabel('Temperature (K)')
        ax.set_ylabel('Mutant kcat (s⁻¹)')
        ax.set_title('Temperature Effect on Catalytic Activity\n(Colored by pH where available)')
        ax.legend()
        ax.grid(True, alpha=0.3)

    def plot_substrate_specificity(self, df, ax):
        """Plot top substrates by specificity constant"""
        valid_data = df.dropna(subset=['substrate_kinetics', 'specificity_constant_mut'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid substrate specificity data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Get top substrates by mean specificity constant
        substrate_stats = valid_data.groupby('substrate_kinetics').agg({
            'specificity_constant_mut': ['mean', 'count', 'std']
        }).round(6)
        
        substrate_stats.columns = ['mean_specificity', 'count', 'std']
        substrate_stats = substrate_stats[substrate_stats['count'] >= 1]  # At least 1 measurement
        top_substrates = substrate_stats.nlargest(15, 'mean_specificity')
        
        if top_substrates.empty:
            ax.text(0.5, 0.5, 'No substrates with sufficient data', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        y_pos = range(len(top_substrates))
        bars = ax.barh(y_pos, top_substrates['mean_specificity'], 
                      xerr=top_substrates['std'], alpha=0.7, edgecolor='black')
        
        # Color bars based on specificity value
        for i, bar in enumerate(bars):
            bar.set_facecolor(plt.cm.viridis(i / len(bars)))
        
        ax.set_yticks(y_pos)
        ax.set_yticklabels([s[:25] + '...' if len(s) > 25 else s for s in top_substrates.index])
        ax.set_xlabel('Specificity Constant (kcat/Km)')
        ax.set_title('Top Substrates by Specificity Constant\n(Bars show mean ± std dev)')
        ax.grid(True, alpha=0.3, axis='x')
        
        # Add value labels
        for i, v in enumerate(top_substrates['mean_specificity']):
            ax.text(v + (v * 0.01), i, f'{v:.2e}', va='center', fontsize=8)

    def plot_efficiency_correlation(self, df, ax):
        """Plot correlation between ΔΔG and catalytic efficiency"""
        valid_data = df.dropna(subset=['ddG', 'catalytic_efficiency_change'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid data for correlation analysis', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create scatter plot
        scatter = ax.scatter(valid_data['ddG'], valid_data['catalytic_efficiency_change'], 
                           alpha=0.6, s=50, c=valid_data.get('temperature', 300), cmap='coolwarm')
        
        # Add reference lines
        ax.axhline(1, color='red', linestyle='--', alpha=0.5, linewidth=2, label='No efficiency change')
        ax.axvline(0, color='blue', linestyle='--', alpha=0.5, linewidth=2, label='Stability threshold')
        
        # Add regression line and statistics
        try:
            z = np.polyfit(valid_data['ddG'], valid_data['catalytic_efficiency_change'], 1)
            p = np.poly1d(z)
            ax.plot(valid_data['ddG'], p(valid_data['ddG']), "k-", alpha=0.8, linewidth=2, label='Regression line')
            
            # Calculate correlation
            r_value = np.corrcoef(valid_data['ddG'], valid_data['catalytic_efficiency_change'])[0,1]
            ax.text(0.02, 0.98, f'Pearson r = {r_value:.3f}\nSlope = {z[0]:.3f}', 
                   transform=ax.transAxes, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="white"),
                   verticalalignment='top')
        except:
            pass
        
        if 'temperature' in valid_data.columns and not valid_data['temperature'].isna().all():
            plt.colorbar(scatter, ax=ax, label='Temperature (K)')
        
        ax.set_xlabel('ΔΔG (kcal/mol)')
        ax.set_ylabel('Catalytic Efficiency Change (mut/wt)')
        ax.set_title('Correlation: Stability vs Catalytic Efficiency\n(Colored by temperature)')
        ax.legend()
        ax.grid(True, alpha=0.3)

    def plot_mutation_impact(self, df, ax):
        """Plot impact of different mutation types on kinetics"""
        if 'mutation_category' not in df.columns:
            ax.text(0.5, 0.5, 'Mutation type data not available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        valid_data = df.dropna(subset=['mutation_category', 'kcat_ratio'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid mutation impact data', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Group by mutation category
        impact_data = valid_data.groupby('mutation_category').agg({
            'kcat_ratio': ['mean', 'std', 'count'],
            'Km_ratio': 'mean',
            'ddG': 'mean'
        }).round(4)
        
        impact_data.columns = ['kcat_ratio_mean', 'kcat_ratio_std', 'count', 'km_ratio_mean', 'ddg_mean']
        impact_data = impact_data[impact_data['count'] >= 3]  # Only categories with sufficient data
        impact_data = impact_data.sort_values('kcat_ratio_mean')
        
        if impact_data.empty:
            ax.text(0.5, 0.5, 'No mutation categories with sufficient data', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        y_pos = range(len(impact_data))
        
        # Create horizontal bar plot
        bars = ax.barh(y_pos, impact_data['kcat_ratio_mean'], 
                      xerr=impact_data['kcat_ratio_std'], alpha=0.7, edgecolor='black')
        
        # Color bars based on effect
        for i, (bar, ratio) in enumerate(zip(bars, impact_data['kcat_ratio_mean'])):
            if ratio > 1.2:
                bar.set_facecolor('green')  # Strong improvement
            elif ratio > 1.0:
                bar.set_facecolor('lightgreen')  # Mild improvement
            elif ratio < 0.8:
                bar.set_facecolor('red')  # Strong reduction
            elif ratio < 1.0:
                bar.set_facecolor('lightcoral')  # Mild reduction
            else:
                bar.set_facecolor('gray')  # No change
        
        ax.axvline(1, color='black', linestyle='--', alpha=0.5, linewidth=2, label='No change')
        
        ax.set_yticks(y_pos)
        ax.set_yticklabels(impact_data.index)
        ax.set_xlabel('Average kcat Ratio (mut/wt)')
        ax.set_title('Impact of Mutation Types on Catalytic Activity\n(Error bars show standard deviation)')
        ax.legend()
        ax.grid(True, alpha=0.3, axis='x')
        
        # Add sample size annotations
        for i, (idx, row) in enumerate(impact_data.iterrows()):
            ax.text(row['kcat_ratio_mean'] + 0.1, i, f'n={row["count"]}', va='center', fontsize=8)

    def plot_enzyme_class_comparison(self, df, ax):
        """Compare kinetic parameters across enzyme classes"""
        if 'ec_class' not in df.columns:
            ax.text(0.5, 0.5, 'EC class data not available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        valid_data = df.dropna(subset=['ec_class', 'specificity_constant_mut'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid data for enzyme class comparison', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Get enzyme classes with sufficient data
        class_counts = valid_data['ec_class'].value_counts()
        significant_classes = class_counts[class_counts >= 5].index  # At least 5 entries
        
        if len(significant_classes) == 0:
            ax.text(0.5, 0.5, 'No enzyme classes with sufficient data', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        filtered_data = valid_data[valid_data['ec_class'].isin(significant_classes)]
        
        # Create boxplot
        box_data = []
        labels = []
        for ec_class in significant_classes:
            class_data = filtered_data[filtered_data['ec_class'] == ec_class]['specificity_constant_mut']
            if len(class_data) > 0:
                box_data.append(class_data)
                labels.append(f'EC {ec_class}\n(n={len(class_data)})')
        
        box_plot = ax.boxplot(box_data, labels=labels, patch_artist=True)
        
        # Color boxes
        colors = plt.cm.Set3(np.linspace(0, 1, len(box_data)))
        for patch, color in zip(box_plot['boxes'], colors):
            patch.set_facecolor(color)
        
        ax.set_ylabel('Specificity Constant (kcat/Km)')
        ax.set_title('Specificity Constant Distribution by Enzyme Class')
        ax.grid(True, alpha=0.3, axis='y')
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')

    def plot_ph_effect(self, df, ax):
        """Plot effect of pH on enzyme activity"""
        valid_data = df.dropna(subset=['pH', 'kcat_mut'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid pH-kcat data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Group by pH and calculate statistics
        ph_stats = valid_data.groupby('pH').agg({
            'kcat_mut': ['mean', 'std', 'count'],
            'Km_mut': 'mean'
        }).round(4)
        
        ph_stats.columns = ['kcat_mean', 'kcat_std', 'count', 'km_mean']
        ph_stats = ph_stats[ph_stats['count'] >= 2]  # Only pH with multiple measurements
        
        if ph_stats.empty:
            ax.text(0.5, 0.5, 'No pH values with sufficient data', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create line plot with error bars
        x = ph_stats.index
        y = ph_stats['kcat_mean']
        y_err = ph_stats['kcat_std']
        
        ax.errorbar(x, y, yerr=y_err, fmt='o-', linewidth=2, markersize=8, 
                   capsize=5, capthick=2, alpha=0.7, label='kcat ± std')
        
        ax.set_xlabel('pH')
        ax.set_ylabel('kcat (s⁻¹)')
        ax.set_title('Effect of pH on Catalytic Activity\n(Error bars show standard deviation)')
        ax.grid(True, alpha=0.3)
        ax.legend()
        
        # Add optimal pH annotation
        optimal_ph = ph_stats.loc[ph_stats['kcat_mean'].idxmax()]
        ax.axvline(ph_stats['kcat_mean'].idxmax(), color='red', linestyle='--', alpha=0.7,
                  label=f'Optimal pH: {ph_stats["kcat_mean"].idxmax()}')

    def plot_kcat_km_ratio(self, df, ax):
        """Plot distribution of kcat/Km ratios"""
        valid_data = df.dropna(subset=['specificity_constant_mut'])
        
        if valid_data.empty:
            ax.text(0.5, 0.5, 'No valid kcat/Km data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create histogram on log scale
        data = np.log10(valid_data['specificity_constant_mut'].clip(lower=1e-10))
        
        ax.hist(data, bins=30, alpha=0.7, edgecolor='black', density=True)
        
        # Add distribution curve
        try:
            from scipy.stats import gaussian_kde
            kde = gaussian_kde(data)
            x_range = np.linspace(data.min(), data.max(), 100)
            ax.plot(x_range, kde(x_range), 'r-', linewidth=2, label='Density')
        except:
            pass
        
        ax.set_xlabel('log10(Specificity Constant) [log10(kcat/Km)]')
        ax.set_ylabel('Density')
        ax.set_title('Distribution of Catalytic Efficiency (kcat/Km)\n(Logarithmic scale)')
        ax.grid(True, alpha=0.3)
        
        # Add statistics
        original_data = valid_data['specificity_constant_mut']
        ax.text(0.02, 0.98, f'Mean: {original_data.mean():.2e}\n'
                           f'Median: {original_data.median():.2e}\n'
                           f'Range: {original_data.min():.2e} - {original_data.max():.2e}',
               transform=ax.transAxes, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="white"),
               verticalalignment='top')

    def plot_correlation_matrix(self, df, ax):
        """Plot correlation matrix between parameters"""
        numeric_cols = ['kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG', 'temperature', 'pH', 
                       'kcat_ratio', 'Km_ratio', 'specificity_constant_mut']
        numeric_cols = [col for col in numeric_cols if col in df.columns]
        
        if len(numeric_cols) < 2:
            ax.text(0.5, 0.5, 'Not enough numeric columns for correlation matrix', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Calculate correlation matrix
        corr_matrix = df[numeric_cols].corr()
        
        # Create heatmap
        im = ax.imshow(corr_matrix, cmap='coolwarm', aspect='auto', vmin=-1, vmax=1)
        
        # Add colorbar
        plt.colorbar(im, ax=ax, label='Correlation Coefficient')
        
        # Add text annotations
        for i in range(len(corr_matrix)):
            for j in range(len(corr_matrix)):
                text = ax.text(j, i, f'{corr_matrix.iloc[i, j]:.2f}',
                              ha="center", va="center", color="black" if abs(corr_matrix.iloc[i, j]) < 0.5 else "white",
                              fontsize=8)
        
        ax.set_xticks(range(len(corr_matrix)))
        ax.set_yticks(range(len(corr_matrix)))
        ax.set_xticklabels([col.replace('_', ' ').title() for col in corr_matrix.columns], rotation=45, ha='right')
        ax.set_yticklabels([col.replace('_', ' ').title() for col in corr_matrix.columns])
        ax.set_title('Correlation Matrix of Kinetic Parameters')

    def plot_kinetic_parameters_distribution(self, df, ax):
        """Plot distribution of multiple kinetic parameters"""
        parameters = ['kcat_mut', 'Km_mut', 'ddG', 'kcat_ratio', 'Km_ratio']
        available_params = [p for p in parameters if p in df.columns and not df[p].isna().all()]
        
        if len(available_params) == 0:
            ax.text(0.5, 0.5, 'No kinetic parameters data available', 
                   ha='center', va='center', transform=ax.transAxes)
            return
        
        # Create subplots within the main axis
        n_params = len(available_params)
        fig, axes = plt.subplots(1, n_params, figsize=(5*n_params, 6))
        
        if n_params == 1:
            axes = [axes]
        
        for i, param in enumerate(available_params):
            valid_data = df[param].dropna()
            
            if param in ['kcat_ratio', 'Km_ratio']:
                # Use log scale for ratios
                data = np.log10(valid_data.clip(lower=1e-10))
                axes[i].hist(data, bins=20, alpha=0.7, edgecolor='black')
                axes[i].set_xlabel(f'log10({param.replace("_", " ")})')
            else:
                axes[i].hist(valid_data, bins=20, alpha=0.7, edgecolor='black')
                axes[i].set_xlabel(param.replace('_', ' ').title())
            
            axes[i].set_ylabel('Frequency')
            axes[i].set_title(f'Distribution of {param.replace("_", " ").title()}')
            axes[i].grid(True, alpha=0.3)
            
            # Add statistics
            stats_text = f'Mean: {valid_data.mean():.3f}\nStd: {valid_data.std():.3f}\nn: {len(valid_data)}'
            axes[i].text(0.05, 0.95, stats_text, transform=axes[i].transAxes, fontsize=9,
                        bbox=dict(boxstyle="round,pad=0.3", facecolor="white"),
                        verticalalignment='top')
        
        plt.tight_layout()

    def setup_enzyme_info_tab(self):
        """Setup the enzyme information tab with comprehensive layout"""
        # Main frame
        main_frame = ttk.Frame(self.enzyme_info_tab)
        main_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Top frame - Basic information
        top_frame = ttk.LabelFrame(main_frame, text="Basic Enzyme Information", padding=10)
        top_frame.pack(fill='x', padx=5, pady=5)
        
        self.basic_info_text = scrolledtext.ScrolledText(top_frame, wrap=tk.WORD, width=100, height=8,
                                                        font=('Consolas', 10))
        self.basic_info_text.pack(fill='both', expand=True)
        
        # Middle frame - Kinetic parameters (two columns)
        middle_frame = ttk.Frame(main_frame)
        middle_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Left column - Wild-type parameters
        left_frame = ttk.LabelFrame(middle_frame, text="Wild-type Parameters", padding=10)
        left_frame.pack(side='left', fill='both', expand=True, padx=5, pady=5)
        
        self.wt_info_text = scrolledtext.ScrolledText(left_frame, wrap=tk.WORD, width=45, height=12,
                                                     font=('Consolas', 10))
        self.wt_info_text.pack(fill='both', expand=True)
        
        # Right column - Mutant parameters
        right_frame = ttk.LabelFrame(middle_frame, text="Mutant Parameters & Changes", padding=10)
        right_frame.pack(side='right', fill='both', expand=True, padx=5, pady=5)
        
        self.mutant_info_text = scrolledtext.ScrolledText(right_frame, wrap=tk.WORD, width=45, height=12,
                                                         font=('Consolas', 10))
        self.mutant_info_text.pack(fill='both', expand=True)
        
        # Bottom frame - Additional details
        bottom_frame = ttk.LabelFrame(main_frame, text="Additional Details & Analysis", padding=10)
        bottom_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.additional_info_text = scrolledtext.ScrolledText(bottom_frame, wrap=tk.WORD, width=100, height=8,
                                                             font=('Consolas', 10))
        self.additional_info_text.pack(fill='both', expand=True)
        
        # Button frame
        button_frame = ttk.Frame(self.enzyme_info_tab)
        button_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(button_frame, text="📊 Generate Specific Plots", 
                  command=self.generate_specific_plots).pack(side='left', padx=5)
        ttk.Button(button_frame, text="💾 Export Enzyme Data", 
                  command=self.export_enzyme_data).pack(side='left', padx=5)
        ttk.Button(button_frame, text="🔄 Refresh Display", 
                  command=self.refresh_display).pack(side='left', padx=5)

    def display_detailed_enzyme_info(self, row):
        """Display detailed enzyme information across multiple sections"""
        
        # --- FIX START ---
        # Set widgets to NORMAL state so they can be cleared
        for widget in [self.basic_info_text, self.wt_info_text, self.mutant_info_text, self.additional_info_text]:
            widget.config(state=tk.NORMAL)
        # --- FIX END ---
        
        # Clear previous content
        self.basic_info_text.delete(1.0, tk.END)
        self.wt_info_text.delete(1.0, tk.END)
        self.mutant_info_text.delete(1.0, tk.END)
        self.additional_info_text.delete(1.0, tk.END)
        
        try:
            # Basic information
            basic_info = f"""
EC Number:          {row.get('ec_number', 'N/A')}
UniProt ID:         {row.get('uniprot', 'N/A')}
Organism:           {row.get('organism', 'N/A')}
Substrate:          {row.get('substrate_kinetics', 'N/A')}
Mutation:           {row.get('clean_mut_wt', 'N/A')}

Amino Acid Change:
  Wild-type:        {row.get('wt_aa', 'N/A')} ({row.get('wt_aa_type', 'N/A')})
  Mutant:           {row.get('mut_aa', 'N/A')} ({row.get('mut_aa_type', 'N/A')})
  Position:         {row.get('position', 'N/A')}

Experimental Conditions:
  Temperature:      {row.get('temperature', 'N/A')} K
  pH:               {row.get('pH', 'N/A')}

Database Information:
  Source:           {row.get('source', 'N/A')}
  Reference:        {row.get('reference', 'N/A')}
"""
            
            # Wild-type parameters
            wt_info = f"""
CATALYTIC PARAMETERS:
  kcat:             {self.format_value(row.get('kcat_wt'), 4)} s⁻¹
  Km:               {self.format_value(row.get('Km_wt'), 6)} M
  Specificity:      {self.format_value(row.get('specificity_constant_wt'), 4)} M⁻¹s⁻¹

THERMODYNAMIC PARAMETERS:
  ΔG:               {self.format_value(row.get('dG_wt'), 4)} kcal/mol

EXPERIMENTAL CONDITIONS:
  Temperature:      {self.format_value(row.get('temperature'), 1)} K
  pH:               {self.format_value(row.get('pH'), 1)}
  
MEASUREMENT QUALITY:
  Data Points:      {row.get('data_points', 'N/A')}
  Method:           {row.get('method', 'N/A')}
"""
            
            # Mutant parameters and changes
            kcat_ratio = row.get('kcat_ratio', 'N/A')
            km_ratio = row.get('Km_ratio', 'N/A')
            efficiency_change = row.get('catalytic_efficiency_change', 'N/A')
            ddg = row.get('ddG', 'N/A')
            
            mutant_info = f"""
CATALYTIC PARAMETERS:
  kcat:             {self.format_value(row.get('kcat_mut'), 4)} s⁻¹
  Km:               {self.format_value(row.get('Km_mut'), 6)} M
  Specificity:      {self.format_value(row.get('specificity_constant_mut'), 4)} M⁻¹s⁻¹

PARAMETER CHANGES:
  kcat Ratio:       {self.format_value(kcat_ratio, 4)} {self.get_change_arrow(kcat_ratio, 'kcat')}
  Km Ratio:         {self.format_value(km_ratio, 4)} {self.get_change_arrow(km_ratio, 'km')}
  Efficiency:       {self.format_value(efficiency_change, 4)} {self.get_change_arrow(efficiency_change, 'efficiency')}

THERMODYNAMIC CHANGES:
  ΔΔG:              {self.format_value(ddg, 4)} kcal/mol {self.get_stability_effect(ddg)}
  ΔG mutant:        {self.format_value(row.get('dG_mut'), 4)} kcal/mol

INTERPRETATION:
  Catalytic Effect: {self.interpret_catalytic_effect(kcat_ratio, km_ratio)}
  Stability Effect: {self.interpret_stability_effect(ddg)}
"""
            
            # Additional details
            additional_info = f"""
MUTATION IMPACT ANALYSIS:
  Stability:        {self.get_impact_level(ddg, 'stability')}
  Catalysis:        {self.get_impact_level(kcat_ratio, 'catalysis')}
  Binding:          {self.get_impact_level(km_ratio, 'binding')}
  Overall:          {self.get_overall_impact(kcat_ratio, km_ratio, ddg)}

STRUCTURAL INFORMATION:
  Domain:           {row.get('domain', 'N/A')}
  Active Site:      {row.get('active_site', 'N/A')}
  Secondary Str:    {row.get('secondary_structure', 'N/A')}

EXTERNAL DATABASES:
  UniProt:          {row.get('uniprot_website', 'N/A')}
  PDB:              {row.get('pdb_id', 'N/A')}
  BRENDA:           {row.get('brenda_id', 'N/A')}

ADDITIONAL NOTES:
  {row.get('notes', 'No additional notes available')}
"""
            
            self.basic_info_text.insert(1.0, basic_info)
            self.wt_info_text.insert(1.0, wt_info)
            self.mutant_info_text.insert(1.0, mutant_info)
            self.additional_info_text.insert(1.0, additional_info)
            
            # Make text widgets read-only
            for widget in [self.basic_info_text, self.wt_info_text, self.mutant_info_text, self.additional_info_text]:
                widget.config(state=tk.DISABLED)
                
        except Exception as e:
            messagebox.showerror("Display Error", f"Failed to display enzyme details: {str(e)}")

    def format_value(self, value, decimals):
        """Format numeric values with proper handling of None/NaN"""
        if pd.isna(value) or value is None:
            return "N/A"
        try:
            return f"{float(value):.{decimals}f}"
        except (ValueError, TypeError):
            return "N/A"

    def get_change_arrow(self, ratio, param_type):
        """Get arrow symbol indicating change direction"""
        if pd.isna(ratio) or ratio is None:
            return ""
        
        try:
            ratio = float(ratio)
            if param_type == 'km':  # For Km, lower is better
                if ratio < 0.8:
                    return "↓"  # Improved binding
                elif ratio > 1.2:
                    return "↑"  # Worse binding
                else:
                    return "→"  # No significant change
            else:  # For kcat and efficiency, higher is better
                if ratio > 1.2:
                    return "↑"  # Improved
                elif ratio < 0.8:
                    return "↓"  # Worse
                else:
                    return "→"  # No significant change
        except (ValueError, TypeError):
            return ""

    def get_stability_effect(self, ddg):
        """Get stability effect symbol"""
        if pd.isna(ddg) or ddg is None:
            return ""
        
        try:
            ddg = float(ddg)
            if ddg < -0.5:
                return "↑"  # Stabilizing
            elif ddg > 0.5:
                return "↓"  # Destabilizing
            else:
                return "→"  # Neutral
        except (ValueError, TypeError):
            return ""

    def interpret_catalytic_effect(self, kcat_ratio, km_ratio):
        """Interpret the overall catalytic effect"""
        if pd.isna(kcat_ratio) or pd.isna(km_ratio):
            return "Insufficient data"
        
        try:
            kcat_effect = "improved" if kcat_ratio > 1.2 else "reduced" if kcat_ratio < 0.8 else "similar"
            km_effect = "improved" if km_ratio < 0.8 else "reduced" if km_ratio > 1.2 else "similar"
            
            if kcat_effect == "improved" and km_effect == "improved":
                return "Strong improvement in both turnover and binding"
            elif kcat_effect == "improved":
                return "Improved turnover with similar/binding"
            elif km_effect == "improved":
                return "Improved binding with similar turnover"
            elif kcat_effect == "reduced" and km_effect == "reduced":
                return "Reduced both turnover and binding"
            else:
                return "Mixed or neutral effects"
        except (ValueError, TypeError):
            return "Cannot interpret"

    def interpret_stability_effect(self, ddg):
        """Interpret stability effect"""
        if pd.isna(ddg):
            return "Unknown stability effect"
        
        try:
            ddg = float(ddg)
            if ddg < -1.0:
                return "Strongly stabilizing"
            elif ddg < -0.5:
                return "Moderately stabilizing"
            elif ddg < 0.5:
                return "Minimal stability effect"
            elif ddg < 1.0:
                return "Moderately destabilizing"
            else:
                return "Strongly destabilizing"
        except (ValueError, TypeError):
            return "Unknown stability effect"

    def get_impact_level(self, value, impact_type):
        """Get impact level description"""
        if pd.isna(value):
            return "Unknown"
        
        try:
            value = float(value)
            if impact_type == 'stability':
                if value < -1.0: return "High (Stabilizing)"
                elif value < -0.5: return "Medium (Stabilizing)"
                elif value < 0.5: return "Low"
                elif value < 1.0: return "Medium (Destabilizing)"
                else: return "High (Destabilizing)"
            elif impact_type == 'catalysis':
                if value > 2.0: return "Very High (Improved)"
                elif value > 1.5: return "High (Improved)"
                elif value > 1.2: return "Medium (Improved)"
                elif value < 0.5: return "High (Reduced)"
                elif value < 0.8: return "Medium (Reduced)"
                else: return "Low"
            elif impact_type == 'binding':
                if value < 0.5: return "High (Improved)"
                elif value < 0.8: return "Medium (Improved)"
                elif value > 2.0: return "High (Reduced)"
                elif value > 1.5: return "Medium (Reduced)"
                else: return "Low"
        except (ValueError, TypeError):
            return "Unknown"

    def get_overall_impact(self, kcat_ratio, km_ratio, ddg):
        """Get overall impact assessment"""
        try:
            impacts = []
            if not pd.isna(kcat_ratio):
                if kcat_ratio > 1.2: impacts.append("Improved catalysis")
                elif kcat_ratio < 0.8: impacts.append("Reduced catalysis")
            
            if not pd.isna(km_ratio):
                if km_ratio < 0.8: impacts.append("Improved binding")
                elif km_ratio > 1.2: impacts.append("Reduced binding")
            
            if not pd.isna(ddg):
                if ddg < -0.5: impacts.append("Stabilizing")
                elif ddg > 0.5: impacts.append("Destabilizing")
            
            if not impacts:
                return "Neutral or unknown impact"
            
            return " | ".join(impacts)
        except (ValueError, TypeError):
            return "Unknown overall impact"

    def generate_specific_plots(self):
        """Generate plots specific to the current enzyme"""
        if self.current_enzyme_data is None:
            messagebox.showwarning("No Selection", "Please select an enzyme first")
            return
        
        # This would open a dialog to select which specific plots to generate
        # For now, switch to visualization tab
        self.notebook.select(1)  # Visualization tab
        messagebox.showinfo("Specific Plots", "Switch to Visualization tab and select plot type for detailed analysis")

    def export_enzyme_data(self):
        """Export current enzyme data"""
        if self.current_enzyme_data is None:
            messagebox.showwarning("No Selection", "Please select an enzyme first")
            return
        
        try:
            # Convert series to dataframe and export
            export_df = pd.DataFrame([self.current_enzyme_data])
            export_df.to_csv('selected_enzyme_data.csv', index=False)
            messagebox.showinfo("Export Successful", "Enzyme data exported to 'selected_enzyme_data.csv'")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    def refresh_display(self):
        """Refresh the current display"""
        if self.current_enzyme_data is not None:
            self.display_detailed_enzyme_info(self.current_enzyme_data)
            self.status_var.set("Display refreshed")

    def setup_data_entry_tab(self):
        """Setup the data entry tab for adding new enzyme data"""
        main_frame = ttk.Frame(self.data_entry_tab)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
    
        # Create a form for data entry
        form_frame = ttk.LabelFrame(main_frame, text="Add New Enzyme Data", padding=15)
        form_frame.pack(fill='both', expand=True)
    
        # Basic information
        basic_frame = ttk.LabelFrame(form_frame, text="Basic Information", padding=10)
        basic_frame.pack(fill='x', padx=5, pady=5)
    
        # EC Number
        ttk.Label(basic_frame, text="EC Number:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.ec_entry = ttk.Entry(basic_frame, width=20)
        self.ec_entry.grid(row=0, column=1, padx=5, pady=2)
    
        # Organism
        ttk.Label(basic_frame, text="Organism:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.org_entry = ttk.Entry(basic_frame, width=20)
        self.org_entry.grid(row=0, column=3, padx=5, pady=2)
    
        # Substrate
        ttk.Label(basic_frame, text="Substrate:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.substrate_entry = ttk.Entry(basic_frame, width=20)
        self.substrate_entry.grid(row=1, column=1, padx=5, pady=2)
    
        # Mutation
        ttk.Label(basic_frame, text="Mutation:").grid(row=1, column=2, sticky='w', padx=5, pady=2)
        self.mutation_entry = ttk.Entry(basic_frame, width=20)
        self.mutation_entry.grid(row=1, column=3, padx=5, pady=2)
    
        # Wild-type parameters
        wt_frame = ttk.LabelFrame(form_frame, text="Wild-type Parameters", padding=10)
        wt_frame.pack(fill='x', padx=5, pady=5)
    
        ttk.Label(wt_frame, text="kcat (s⁻¹):").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.kcat_wt_entry = ttk.Entry(wt_frame, width=15)
        self.kcat_wt_entry.grid(row=0, column=1, padx=5, pady=2)
    
        ttk.Label(wt_frame, text="Km (M):").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.km_wt_entry = ttk.Entry(wt_frame, width=15)
        self.km_wt_entry.grid(row=0, column=3, padx=5, pady=2)
    
        # Mutant parameters
        mut_frame = ttk.LabelFrame(form_frame, text="Mutant Parameters", padding=10)
        mut_frame.pack(fill='x', padx=5, pady=5)
    
        ttk.Label(mut_frame, text="kcat (s⁻¹):").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.kcat_mut_entry = ttk.Entry(mut_frame, width=15)
        self.kcat_mut_entry.grid(row=0, column=1, padx=5, pady=2)
    
        ttk.Label(mut_frame, text="Km (M):").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.km_mut_entry = ttk.Entry(mut_frame, width=15)
        self.km_mut_entry.grid(row=0, column=3, padx=5, pady=2)
    
        ttk.Label(mut_frame, text="ΔΔG (kcal/mol):").grid(row=0, column=4, sticky='w', padx=5, pady=2)
        self.ddg_entry = ttk.Entry(mut_frame, width=15)
        self.ddg_entry.grid(row=0, column=5, padx=5, pady=2)
    
        # Experimental conditions
        exp_frame = ttk.LabelFrame(form_frame, text="Experimental Conditions", padding=10)
        exp_frame.pack(fill='x', padx=5, pady=5)
    
        ttk.Label(exp_frame, text="Temperature (K):").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.temp_entry = ttk.Entry(exp_frame, width=15)
        self.temp_entry.grid(row=0, column=1, padx=5, pady=2)
    
        ttk.Label(exp_frame, text="pH:").grid(row=0, column=2, sticky='w', padx=5, pady=2)
        self.ph_entry = ttk.Entry(exp_frame, width=15)
        self.ph_entry.grid(row=0, column=3, padx=5, pady=2)
    
        # Buttons - ADDED REMOVE BUTTON HERE
        button_frame = ttk.Frame(form_frame)
        button_frame.pack(fill='x', padx=5, pady=10)
    
        ttk.Button(button_frame, text="➕ Add Data", 
              command=self.add_new_data).pack(side='left', padx=5)
        ttk.Button(button_frame, text="🗑️ Clear Form", 
              command=self.clear_data_form).pack(side='left', padx=5)
        ttk.Button(button_frame, text="🧮 Calculate Ratios", 
              command=self.calculate_ratios).pack(side='left', padx=5)
        # NEW: Remove last added data button
        ttk.Button(button_frame, text="❌ Remove Last Added", 
              command=self.remove_last_added_data).pack(side='left', padx=5)
    
        # Results display
        results_frame = ttk.LabelFrame(main_frame, text="Calculated Parameters", padding=10)
        results_frame.pack(fill='x', padx=5, pady=5)
    
        self.results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, width=100, height=8,
                                                 font=('Consolas', 9))
        self.results_text.pack(fill='both', expand=True)
        self.results_text.config(state=tk.DISABLED)

    def add_new_data(self):
        """Add new data to the dataset"""
        try:
            # Get values from form
            new_data = {
                'ec_number': self.ec_entry.get(),
                'organism': self.org_entry.get(),
                'substrate_kinetics': self.substrate_entry.get(),
                'clean_mut_wt': self.mutation_entry.get(),
                'kcat_wt': float(self.kcat_wt_entry.get()) if self.kcat_wt_entry.get() else None,
                'Km_wt': float(self.km_wt_entry.get()) if self.km_wt_entry.get() else None,
                'kcat_mut': float(self.kcat_mut_entry.get()) if self.kcat_mut_entry.get() else None,
                'Km_mut': float(self.km_mut_entry.get()) if self.km_mut_entry.get() else None,
                'ddG': float(self.ddg_entry.get()) if self.ddg_entry.get() else None,
                'temperature': float(self.temp_entry.get()) if self.temp_entry.get() else None,
                'pH': float(self.ph_entry.get()) if self.ph_entry.get() else None
            }
            
            # Validate required fields
            if not new_data['ec_number'] or not new_data['organism']:
                messagebox.showwarning("Input Error", "EC Number and Organism are required fields")
                return
            
            # Add to dataframe
            new_row = pd.DataFrame([new_data])
            self.df = pd.concat([self.df, new_row], ignore_index=True)
            
            # Re-preprocess data to calculate derived columns
            self.preprocess_data()
            
            messagebox.showinfo("Success", "New data added successfully!")
            self.clear_data_form()
            self.status_var.set(f"New data added | Total entries: {len(self.df):,}")
            
        except ValueError as e:
            messagebox.showerror("Input Error", "Please enter valid numeric values for parameters")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add new data: {str(e)}")

    def clear_data_form(self):
        """Clear the data entry form"""
        self.ec_entry.delete(0, tk.END)
        self.org_entry.delete(0, tk.END)
        self.substrate_entry.delete(0, tk.END)
        self.mutation_entry.delete(0, tk.END)
        self.kcat_wt_entry.delete(0, tk.END)
        self.km_wt_entry.delete(0, tk.END)
        self.kcat_mut_entry.delete(0, tk.END)
        self.km_mut_entry.delete(0, tk.END)
        self.ddg_entry.delete(0, tk.END)
        self.temp_entry.delete(0, tk.END)
        self.ph_entry.delete(0, tk.END)
        
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)

    def calculate_ratios(self):
        """Calculate and display derived parameters"""
        try:
            kcat_wt = float(self.kcat_wt_entry.get()) if self.kcat_wt_entry.get() else 0
            km_wt = float(self.km_wt_entry.get()) if self.km_wt_entry.get() else 0
            kcat_mut = float(self.kcat_mut_entry.get()) if self.kcat_mut_entry.get() else 0
            km_mut = float(self.km_mut_entry.get()) if self.km_mut_entry.get() else 0
            
            results = "CALCULATED PARAMETERS:\n"
            results += "=" * 50 + "\n"
            
            if kcat_wt > 0 and kcat_mut > 0:
                kcat_ratio = kcat_mut / kcat_wt
                results += f"kcat Ratio (mut/wt): {kcat_ratio:.4f}\n"
            
            if km_wt > 0 and km_mut > 0:
                km_ratio = km_mut / km_wt
                results += f"Km Ratio (mut/wt): {km_ratio:.4f}\n"
            
            if kcat_mut > 0 and km_mut > 0:
                specificity_mut = kcat_mut / km_mut
                results += f"Specificity Constant (mut): {specificity_mut:.4f} M⁻¹s⁻¹\n"
            
            if kcat_wt > 0 and km_wt > 0:
                specificity_wt = kcat_wt / km_wt
                results += f"Specificity Constant (wt): {specificity_wt:.4f} M⁻¹s⁻¹\n"
            
            if 'specificity_mut' in locals() and 'specificity_wt' in locals() and specificity_wt > 0:
                efficiency_change = specificity_mut / specificity_wt
                results += f"Catalytic Efficiency Change: {efficiency_change:.4f}\n"
            
            self.results_text.config(state=tk.NORMAL)
            self.results_text.delete(1.0, tk.END)
            self.results_text.insert(1.0, results)
            self.results_text.config(state=tk.DISABLED)
            
        except ValueError:
            messagebox.showwarning("Input Error", "Please enter valid numeric values first")
        except Exception as e:
            messagebox.showerror("Calculation Error", f"Failed to calculate ratios: {str(e)}")

    def remove_last_added_data(self):
        """Remove the last added data entry from the dataset"""
        if len(self.df) == 0:
           messagebox.showwarning("No Data", "No data entries to remove")
           return
    
        # Get the last row for confirmation
        last_row = self.df.iloc[-1]
    
        # Ask for confirmation
        confirm = messagebox.askyesno(
            "Confirm Removal",
            f"Are you sure you want to remove the last added entry?\n\n"
            f"EC Number: {last_row.get('ec_number', 'N/A')}\n"
            f"Organism: {last_row.get('organism', 'N/A')}\n"
            f"Mutation: {last_row.get('clean_mut_wt', 'N/A')}"
        )
    
        if confirm:
            # Remove the last row
            self.df = self.df.iloc[:-1]
        
            # Re-preprocess data to update derived columns
            self.preprocess_data()
        
            # Update database info
            self.update_db_info()
        
            # Clear the data entry form
            self.clear_data_form()
        
            messagebox.showinfo("Success", "Last added entry removed successfully!")
            self.status_var.set(f"Last entry removed | Total entries: {len(self.df):,}")    

    def setup_stats_tab(self):
        """Setup the statistics tab with comprehensive options"""
        # Control frame
        control_frame = ttk.LabelFrame(self.stats_tab, text="Statistical Analysis Options", padding=10)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Button grid
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(fill='x')
        
        ttk.Button(btn_frame, text="📊 Basic Statistics", 
                  command=self.calculate_basic_statistics).grid(row=0, column=0, padx=5, pady=2)
        ttk.Button(btn_frame, text="📈 Advanced Statistics", 
                  command=self.calculate_advanced_statistics).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(btn_frame, text="🔗 Correlation Analysis", 
                  command=self.calculate_correlations).grid(row=0, column=2, padx=5, pady=2)
        ttk.Button(btn_frame, text="📉 Regression Analysis", 
                  command=self.calculate_regression).grid(row=0, column=3, padx=5, pady=2)
        ttk.Button(btn_frame, text="📋 Normality Tests", 
                  command=self.calculate_normality).grid(row=1, column=0, padx=5, pady=2)
        ttk.Button(btn_frame, text="📁 Distribution Analysis", 
                  command=self.calculate_distributions).grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(btn_frame, text="🧮 Group Comparison", 
                  command=self.calculate_group_comparison).grid(row=1, column=2, padx=5, pady=2)
        ttk.Button(btn_frame, text="🗑️ Clear Results", 
                  command=self.clear_stats).grid(row=1, column=3, padx=5, pady=2)
        
        # Stats display frame
        stats_frame = ttk.LabelFrame(self.stats_tab, text="Statistical Results", padding=10)
        stats_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.stats_text = scrolledtext.ScrolledText(stats_frame, wrap=tk.WORD, width=100, height=30,
                                                   font=('Consolas', 9))
        self.stats_text.pack(fill='both', expand=True)

    def calculate_basic_statistics(self):
        """Calculate basic descriptive statistics"""
        stats_text = "BASIC DESCRIPTIVE STATISTICS\n"
        stats_text += "=" * 70 + "\n\n"
        
        parameters = ['kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG', 'kcat_ratio', 'Km_ratio', 
                     'specificity_constant_mut', 'specificity_constant_wt', 'catalytic_efficiency_change']
        
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 0:
                    stats_text += f"{param.upper()}:\n"
                    stats_text += f"  Count:      {len(valid_data):>8}\n"
                    stats_text += f"  Mean:       {valid_data.mean():>12.6f}\n"
                    stats_text += f"  Std:        {valid_data.std():>12.6f}\n"
                    stats_text += f"  Min:        {valid_data.min():>12.6f}\n"
                    stats_text += f"  Max:        {valid_data.max():>12.6f}\n"
                    stats_text += f"  Median:     {valid_data.median():>12.6f}\n"
                    stats_text += f"  Q1:         {valid_data.quantile(0.25):>12.6f}\n"
                    stats_text += f"  Q3:         {valid_data.quantile(0.75):>12.6f}\n"
                    stats_text += f"  Missing:    {self.df[param].isna().sum():>8}\n\n"
        
        self.display_stats(stats_text)

    def calculate_advanced_statistics(self):
        """Calculate advanced statistical analysis"""
        stats_text = "ADVANCED STATISTICAL ANALYSIS\n"
        stats_text += "=" * 70 + "\n\n"
        
        # Coefficient of variation for key parameters
        stats_text += "COEFFICIENT OF VARIATION (CV = Std/Mean):\n"
        parameters = ['kcat_mut', 'Km_mut', 'ddG', 'kcat_ratio', 'Km_ratio']
        
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 0 and valid_data.mean() != 0:
                    cv = valid_data.std() / valid_data.mean()
                    stats_text += f"  {param}: {cv:.4f} ({'High' if cv > 0.5 else 'Medium' if cv > 0.2 else 'Low'} variability)\n"
        stats_text += "\n"
        
        # Skewness and Kurtosis
        stats_text += "DISTRIBUTION SHAPE ANALYSIS:\n"
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 0:
                    skewness = stats.skew(valid_data)
                    kurtosis = stats.kurtosis(valid_data)
                    stats_text += f"  {param} - Skewness: {skewness:.4f}, Kurtosis: {kurtosis:.4f}\n"
        stats_text += "\n"
        
        self.display_stats(stats_text)

    def calculate_correlations(self):
        """Calculate detailed correlation analysis"""
        stats_text = "DETAILED CORRELATION ANALYSIS\n"
        stats_text += "=" * 70 + "\n\n"
        
        # Select numeric columns for correlation
        numeric_cols = ['kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG', 'temperature', 'pH', 
                       'kcat_ratio', 'Km_ratio', 'catalytic_efficiency_change']
        numeric_cols = [col for col in numeric_cols if col in self.df.columns]
        
        if len(numeric_cols) < 2:
            stats_text += "Not enough numeric columns for correlation analysis\n"
            self.display_stats(stats_text)
            return
        
        # Calculate correlations
        corr_matrix = self.df[numeric_cols].corr()
        
        stats_text += "CORRELATION MATRIX:\n\n"
        stats_text += corr_matrix.to_string(float_format="%.4f")
        stats_text += "\n\n"
        
        # Significant correlations
        stats_text += "SIGNIFICANT CORRELATIONS (|r| > 0.3, p < 0.05):\n"
        significant_found = False
        
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                col1 = corr_matrix.columns[i]
                col2 = corr_matrix.columns[j]
                corr_value = corr_matrix.iloc[i, j]
                
                if not pd.isna(corr_value) and abs(corr_value) > 0.3:
                    # Calculate p-value
                    valid_data = self.df[[col1, col2]].dropna()
                    if len(valid_data) > 2:
                        try:
                            _, p_value = stats.pearsonr(valid_data[col1], valid_data[col2])
                            if p_value < 0.05:
                                strength = "Strong" if abs(corr_value) > 0.7 else "Moderate" if abs(corr_value) > 0.5 else "Weak"
                                direction = "Positive" if corr_value > 0 else "Negative"
                                stats_text += f"  {col1} vs {col2}: r = {corr_value:.4f} ({strength} {direction}, p = {p_value:.4f})\n"
                                significant_found = True
                        except:
                            pass
        
        if not significant_found:
            stats_text += "  No significant correlations found\n"
        
        self.display_stats(stats_text)

    def calculate_regression(self):
        """Perform regression analysis"""
        stats_text = "REGRESSION ANALYSIS\n"
        stats_text += "=" * 70 + "\n\n"
        
        # Example: Regression between ddG and kcat_ratio
        if all(col in self.df.columns for col in ['ddG', 'kcat_ratio']):
            valid_data = self.df[['ddG', 'kcat_ratio']].dropna()
            
            if len(valid_data) > 2:
                x = valid_data['ddG']
                y = valid_data['kcat_ratio']
                
                try:
                    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
                    
                    stats_text += "REGRESSION: ΔΔG vs kcat ratio\n"
                    stats_text += f"  Equation:    kcat_ratio = {slope:.6f} * ΔΔG + {intercept:.6f}\n"
                    stats_text += f"  R-squared:   {r_value**2:.6f}\n"
                    stats_text += f"  P-value:     {p_value:.6f}\n"
                    stats_text += f"  Std error:   {std_err:.6f}\n"
                    stats_text += f"  Data points: {len(valid_data)}\n\n"
                    
                    if p_value < 0.05:
                        stats_text += "  The relationship is statistically significant (p < 0.05)\n"
                        if slope > 0:
                            stats_text += "  Positive relationship: Higher ΔΔG (destabilizing) associated with higher kcat ratio\n"
                        else:
                            stats_text += "  Negative relationship: Higher ΔΔG (destabilizing) associated with lower kcat ratio\n"
                    else:
                        stats_text += "  The relationship is not statistically significant (p >= 0.05)\n"
                        
                except Exception as e:
                    stats_text += f"  Regression analysis failed: {str(e)}\n"
            else:
                stats_text += "  Insufficient data for regression analysis\n"
        else:
            stats_text += "  Required columns (ddG, kcat_ratio) not available\n"
        
        self.display_stats(stats_text)

    def calculate_normality(self):
        """Perform normality tests"""
        stats_text = "NORMALITY TESTS (Shapiro-Wilk)\n"
        stats_text += "=" * 70 + "\n\n"
        
        parameters = ['kcat_mut', 'Km_mut', 'ddG', 'kcat_ratio', 'Km_ratio']
        
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 3 and len(valid_data) <= 5000:  # Shapiro-Wilk limit
                    try:
                        stat, p_value = stats.shapiro(valid_data)
                        normal = "Normal" if p_value > 0.05 else "Non-normal"
                        stats_text += f"  {param}: W = {stat:.4f}, p = {p_value:.4f} ({normal})\n"
                    except:
                        stats_text += f"  {param}: Test failed (possibly constant data)\n"
                else:
                    stats_text += f"  {param}: Insufficient data or too large for test (n={len(valid_data)})\n"
        
        stats_text += "\nNote: p > 0.05 suggests normal distribution\n"
        
        self.display_stats(stats_text)

    def calculate_distributions(self):
        """Calculate distribution statistics"""
        stats_text = "DISTRIBUTION ANALYSIS\n"
        stats_text += "=" * 70 + "\n\n"
        
        parameters = ['kcat_mut', 'Km_mut', 'ddG']
        
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 0:
                    stats_text += f"{param.upper()}:\n"
                    stats_text += f"  Range:        {valid_data.max() - valid_data.min():.6f}\n"
                    stats_text += f"  IQR:          {valid_data.quantile(0.75) - valid_data.quantile(0.25):.6f}\n"
                    stats_text += f"  CV:           {(valid_data.std() / valid_data.mean()):.6f}\n" if valid_data.mean() != 0 else "  CV:           N/A (zero mean)\n"
                    
                    # Outliers using IQR method
                    Q1 = valid_data.quantile(0.25)
                    Q3 = valid_data.quantile(0.75)
                    IQR = Q3 - Q1
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 1.5 * IQR
                    outliers = valid_data[(valid_data < lower_bound) | (valid_data > upper_bound)]
                    stats_text += f"  Outliers:     {len(outliers)} ({len(outliers)/len(valid_data)*100:.1f}%)\n\n"
        
        self.display_stats(stats_text)

    def calculate_group_comparison(self):
        """Compare groups (e.g., by EC class)"""
        stats_text = "GROUP COMPARISON (by EC Class)\n"
        stats_text += "=" * 70 + "\n\n"
        
        if 'ec_class' not in self.df.columns or 'kcat_mut' not in self.df.columns:
            stats_text += "Required columns (ec_class, kcat_mut) not available\n"
            self.display_stats(stats_text)
            return
        
        # Get EC classes with sufficient data
        class_stats = self.df.groupby('ec_class').agg({
            'kcat_mut': ['count', 'mean', 'std'],
            'Km_mut': ['mean', 'std'],
            'ddG': ['mean', 'std']
        }).round(6)
        
        if class_stats.empty:
            stats_text += "No sufficient data for group comparison\n"
            self.display_stats(stats_text)
            return
        
        stats_text += class_stats.to_string()
        stats_text += "\n\n"
        
        # ANOVA test if enough groups
        valid_data = self.df[['ec_class', 'kcat_mut']].dropna()
        classes = valid_data['ec_class'].unique()
        
        if len(classes) >= 2:
            groups = [valid_data[valid_data['ec_class'] == ec_class]['kcat_mut'] for ec_class in classes]
            try:
                f_stat, p_value = stats.f_oneway(*groups)
                stats_text += f"ANOVA TEST (kcat_mut across EC classes):\n"
                stats_text += f"  F-statistic: {f_stat:.4f}, p-value: {p_value:.4f}\n"
                if p_value < 0.05:
                    stats_text += "  Significant differences exist between EC classes (p < 0.05)\n"
                else:
                    stats_text += "  No significant differences between EC classes (p >= 0.05)\n"
            except:
                stats_text += "  ANOVA test could not be performed\n"
        
        self.display_stats(stats_text)

    def clear_stats(self):
        """Clear statistics results"""
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete(1.0, tk.END)
        self.stats_text.config(state=tk.DISABLED)
        self.status_var.set("Statistics cleared")

    def display_stats(self, text):
        """Display statistics text"""
        self.stats_text.config(state=tk.NORMAL)
        self.stats_text.delete(1.0, tk.END)
        self.stats_text.insert(1.0, text)
        self.stats_text.config(state=tk.DISABLED)
        self.status_var.set("Statistics calculated")

    def setup_advanced_tab(self):
        """Setup the advanced analysis tab"""
        # Control frame
        control_frame = ttk.LabelFrame(self.advanced_tab, text="Advanced Analysis Tools", padding=10)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        # Button grid
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(fill='x')
        
        ttk.Button(btn_frame, text="🔗 Correlation Matrix", 
                  command=self.show_correlation_matrix).grid(row=0, column=0, padx=5, pady=2)
        ttk.Button(btn_frame, text="🧬 Mutation Impact", 
                  command=self.analyze_mutation_impact).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(btn_frame, text="📊 Enzyme Class Stats", 
                  command=self.compare_enzyme_classes).grid(row=0, column=2, padx=5, pady=2)
        ttk.Button(btn_frame, text="📈 Outlier Detection", 
                  command=self.detect_outliers).grid(row=0, column=3, padx=5, pady=2)
        ttk.Button(btn_frame, text="📋 Data Quality", 
                  command=self.analyze_data_quality).grid(row=1, column=0, padx=5, pady=2)
        ttk.Button(btn_frame, text="📉 Trend Analysis", 
                  command=self.trend_analysis).grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(btn_frame, text="🧮 Multivariate", 
                  command=self.multivariate_analysis).grid(row=1, column=2, padx=5, pady=2)
        ttk.Button(btn_frame, text="🗑️ Clear", 
                  command=self.clear_advanced).grid(row=1, column=3, padx=5, pady=2)
        
        # Results frame
        results_frame = ttk.LabelFrame(self.advanced_tab, text="Advanced Analysis Results", padding=10)
        results_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.advanced_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, width=100, height=30,
                                                      font=('Consolas', 9))
        self.advanced_text.pack(fill='both', expand=True)

    def show_correlation_matrix(self):
        """Display detailed correlation matrix"""
        text = "ADVANCED CORRELATION MATRIX ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        numeric_cols = ['kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG', 'temperature', 'pH', 
                       'kcat_ratio', 'Km_ratio', 'specificity_constant_mut', 'catalytic_efficiency_change']
        numeric_cols = [col for col in numeric_cols if col in self.df.columns]
        
        if len(numeric_cols) < 2:
            text += "Not enough numeric columns for correlation analysis\n"
            self.display_advanced(text)
            return
        
        corr_matrix = self.df[numeric_cols].corr()
        
        text += "CORRELATION MATRIX:\n\n"
        text += corr_matrix.to_string(float_format="%.4f")
        text += "\n\n"
        
        # Find strongest correlations
        text += "STRONGEST CORRELATIONS (|r| > 0.5):\n"
        strong_corrs = []
        
        for i in range(len(corr_matrix.columns)):
            for j in range(i+1, len(corr_matrix.columns)):
                corr_value = corr_matrix.iloc[i, j]
                if not pd.isna(corr_value) and abs(corr_value) > 0.5:
                    strong_corrs.append((corr_matrix.columns[i], corr_matrix.columns[j], corr_value))
        
        # Sort by absolute correlation strength
        strong_corrs.sort(key=lambda x: abs(x[2]), reverse=True)
        
        for col1, col2, corr in strong_corrs:
            strength = "Very Strong" if abs(corr) > 0.8 else "Strong" if abs(corr) > 0.6 else "Moderate"
            direction = "Positive" if corr > 0 else "Negative"
            text += f"  {col1} ↔ {col2}: r = {corr:.4f} ({strength} {direction})\n"
        
        if not strong_corrs:
            text += "  No strong correlations found\n"
        
        self.display_advanced(text)

    def analyze_mutation_impact(self):
        """Analyze mutation impact on enzyme properties"""
        text = "MUTATION IMPACT ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        if 'mutation_category' not in self.df.columns:
            text += "Mutation category data not available\n"
            self.display_advanced(text)
            return
        
        # Group by mutation category
        impact_stats = self.df.groupby('mutation_category').agg({
            'kcat_ratio': ['count', 'mean', 'std'],
            'Km_ratio': ['mean', 'std'],
            'ddG': ['mean', 'std'],
            'catalytic_efficiency_change': ['mean', 'std']
        }).round(4)
        
        if impact_stats.empty:
            text += "No mutation impact data available\n"
            self.display_advanced(text)
            return
        
        text += "IMPACT BY MUTATION TYPE:\n\n"
        text += impact_stats.to_string()
        text += "\n\n"
        
        # Find most beneficial and detrimental mutations
        if 'kcat_ratio' in self.df.columns:
            beneficial = self.df.nlargest(5, 'kcat_ratio')[['mutation_category', 'kcat_ratio', 'Km_ratio', 'ddG']]
            detrimental = self.df.nsmallest(5, 'kcat_ratio')[['mutation_category', 'kcat_ratio', 'Km_ratio', 'ddG']]
            
            text += "TOP 5 MOST BENEFICIAL MUTATIONS (by kcat ratio):\n"
            for _, row in beneficial.iterrows():
                text += f"  {row['mutation_category']}: kcat_ratio={row['kcat_ratio']:.4f}, Km_ratio={row['Km_ratio']:.4f}, ΔΔG={row['ddG']:.4f}\n"
            
            text += "\nTOP 5 MOST DETRIMENTAL MUTATIONS (by kcat ratio):\n"
            for _, row in detrimental.iterrows():
                text += f"  {row['mutation_category']}: kcat_ratio={row['kcat_ratio']:.4f}, Km_ratio={row['Km_ratio']:.4f}, ΔΔG={row['ddG']:.4f}\n"
        
        self.display_advanced(text)

    def compare_enzyme_classes(self):
        """Compare enzyme classes statistically"""
        text = "ENZYME CLASS COMPARISON\n"
        text += "=" * 80 + "\n\n"
        
        if 'ec_class' not in self.df.columns:
            text += "EC class data not available\n"
            self.display_advanced(text)
            return
        
        # Statistics by EC class
        class_stats = self.df.groupby('ec_class').agg({
            'kcat_mut': ['count', 'mean', 'median', 'std'],
            'Km_mut': ['mean', 'median', 'std'],
            'ddG': ['mean', 'std'],
            'specificity_constant_mut': ['mean', 'median', 'std']
        }).round(6)
        
        text += "STATISTICS BY ENZYME CLASS:\n\n"
        text += class_stats.to_string()
        text += "\n\n"
        
        # Kruskal-Wallis test for non-parametric comparison
        if 'kcat_mut' in self.df.columns:
            valid_data = self.df[['ec_class', 'kcat_mut']].dropna()
            classes = valid_data['ec_class'].unique()
            
            if len(classes) >= 2:
                groups = [valid_data[valid_data['ec_class'] == ec_class]['kcat_mut'] for ec_class in classes]
                try:
                    h_stat, p_value = stats.kruskal(*groups)
                    text += f"KRUSKAL-WALLIS TEST (kcat_mut across EC classes):\n"
                    text += f"  H-statistic: {h_stat:.4f}, p-value: {p_value:.4f}\n"
                    if p_value < 0.05:
                        text += "  Significant differences exist between EC classes (p < 0.05)\n"
                    else:
                        text += "  No significant differences between EC classes (p >= 0.05)\n"
                except:
                    text += "  Kruskal-Wallis test could not be performed\n"
        
        self.display_advanced(text)

    def detect_outliers(self):
        """Detect and analyze outliers"""
        text = "OUTLIER DETECTION ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        parameters = ['kcat_mut', 'Km_mut', 'ddG', 'kcat_ratio', 'Km_ratio']
        
        for param in parameters:
            if param in self.df.columns:
                valid_data = self.df[param].dropna()
                if len(valid_data) > 0:
                    Q1 = valid_data.quantile(0.25)
                    Q3 = valid_data.quantile(0.75)
                    IQR = Q3 - Q1
                    lower_bound = Q1 - 1.5 * IQR
                    upper_bound = Q3 + 1.5 * IQR
                    
                    outliers = valid_data[(valid_data < lower_bound) | (valid_data > upper_bound)]
                    
                    text += f"{param.upper()}:\n"
                    text += f"  Total values: {len(valid_data)}\n"
                    text += f"  Outliers:     {len(outliers)} ({len(outliers)/len(valid_data)*100:.1f}%)\n"
                    if len(outliers) > 0:
                        text += f"  Outlier range: {outliers.min():.6f} to {outliers.max():.6f}\n"
                        text += f"  Main range:    {valid_data[(valid_data >= lower_bound) & (valid_data <= upper_bound)].min():.6f} to {valid_data[(valid_data >= lower_bound) & (valid_data <= upper_bound)].max():.6f}\n"
                    text += "\n"
        
        self.display_advanced(text)

    def analyze_data_quality(self):
        """Analyze data quality and completeness"""
        text = "DATA QUALITY ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        text += f"Total entries in dataset: {len(self.df):,}\n\n"
        
        # Column completeness
        text += "DATA COMPLETENESS BY COLUMN:\n"
        important_cols = ['ec_number', 'organism', 'substrate_kinetics', 'clean_mut_wt', 
                         'kcat_mut', 'Km_mut', 'kcat_wt', 'Km_wt', 'ddG']
        
        for col in important_cols:
            if col in self.df.columns:
                non_null = self.df[col].notna().sum()
                percentage = (non_null / len(self.df)) * 100
                text += f"  {col:<20}: {non_null:>6} / {len(self.df):>6} ({percentage:5.1f}%)\n"
        
        text += "\n"
        
        # Data quality issues
        text += "DATA QUALITY ISSUES:\n"
        
        # Check for zeros in critical columns
        if 'kcat_mut' in self.df.columns:
            zero_kcat = (self.df['kcat_mut'] == 0).sum()
            if zero_kcat > 0:
                text += f"  {zero_kcat} entries have kcat_mut = 0 (biologically unlikely)\n"
        
        if 'Km_mut' in self.df.columns:
            zero_km = (self.df['Km_mut'] == 0).sum()
            if zero_km > 0:
                text += f"  {zero_km} entries have Km_mut = 0 (biologically unlikely)\n"
        
        # Check for extreme values
        if 'kcat_ratio' in self.df.columns:
            extreme_high = (self.df['kcat_ratio'] > 100).sum()
            extreme_low = (self.df['kcat_ratio'] < 0.01).sum()
            if extreme_high > 0 or extreme_low > 0:
                text += f"  {extreme_high + extreme_low} entries have extreme kcat ratios (<0.01 or >100)\n"
        
        self.display_advanced(text)

    def trend_analysis(self):
        """Perform trend analysis on the data"""
        text = "TREND ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        # Temporal trends (if date column exists)
        if 'year' in self.df.columns:
            year_stats = self.df.groupby('year').agg({
                'kcat_mut': 'mean',
                'Km_mut': 'mean',
                'ddG': 'mean'
            })
            text += "TEMPORAL TRENDS (by year):\n"
            text += year_stats.to_string()
            text += "\n\n"
        
        # Size-based trends
        if 'protein_length' in self.df.columns:
            # Bin by protein length
            self.df['length_bin'] = pd.cut(self.df['protein_length'], bins=5)
            length_stats = self.df.groupby('length_bin').agg({
                'kcat_mut': 'mean',
                'Km_mut': 'mean',
                'ddG': 'mean'
            })
            text += "TRENDS BY PROTEIN LENGTH:\n"
            text += length_stats.to_string()
            text += "\n\n"
        
        # pH trends
        if 'pH' in self.df.columns:
            ph_stats = self.df.groupby('pH').agg({
                'kcat_mut': ['mean', 'count'],
                'Km_mut': 'mean'
            })
            text += "TRENDS BY pH:\n"
            text += ph_stats.to_string()
            text += "\n\n"
        
        self.display_advanced(text)

    def multivariate_analysis(self):
        """Perform multivariate analysis"""
        text = "MULTIVARIATE ANALYSIS\n"
        text += "=" * 80 + "\n\n"
        
        text += "Principal Component Analysis and other multivariate methods\n"
        text += "would be implemented here with scikit-learn.\n\n"
        
        text += "Available for future implementation:\n"
        text += "- Principal Component Analysis (PCA)\n"
        text += "- Cluster Analysis\n"
        text += "- Factor Analysis\n"
        text += "- Multidimensional Scaling\n"
        
        self.display_advanced(text)

    def clear_advanced(self):
        """Clear advanced analysis results"""
        self.advanced_text.config(state=tk.NORMAL)
        self.advanced_text.delete(1.0, tk.END)
        self.advanced_text.config(state=tk.DISABLED)
        self.status_var.set("Advanced analysis cleared")

    def display_advanced(self, text):
        """Display advanced analysis text"""
        self.advanced_text.config(state=tk.NORMAL)
        self.advanced_text.delete(1.0, tk.END)
        self.advanced_text.insert(1.0, text)
        self.advanced_text.config(state=tk.DISABLED)
        self.status_var.set("Advanced analysis completed")

    def setup_database_tab(self):
        """Setup the database explorer tab"""
        # Control frame
        control_frame = ttk.LabelFrame(self.database_tab, text="Database Controls", padding=10)
        control_frame.pack(fill='x', padx=10, pady=5)
        
        ttk.Button(control_frame, text="📁 Load All Data", 
                  command=self.load_all_data).pack(side='left', padx=5)
        ttk.Button(control_frame, text="🔍 Quick Filter", 
                  command=self.quick_filter).pack(side='left', padx=5)
        ttk.Button(control_frame, text="💾 Export All Data", 
                  command=self.export_all_data).pack(side='left', padx=5)
        
        # Database info
        info_frame = ttk.Frame(control_frame)
        info_frame.pack(side='right', padx=10)
        
        self.db_info_var = tk.StringVar()
        self.update_db_info()
        ttk.Label(info_frame, textvariable=self.db_info_var, 
                 font=('Arial', 9)).pack()
        
        # Treeview for database
        tree_frame = ttk.Frame(self.database_tab)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        columns = ('ec_number', 'organism', 'substrate', 'mutation', 'kcat_mut', 'Km_mut', 'ddG')
        self.db_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=25)
        
        # Define headings
        headings = {
            'ec_number': 'EC Number',
            'organism': 'Organism', 
            'substrate': 'Substrate',
            'mutation': 'Mutation',
            'kcat_mut': 'kcat (mut)',
            'Km_mut': 'Km (mut)',
            'ddG': 'ΔΔG'
        }
        
        for col, text in headings.items():
            self.db_tree.heading(col, text=text)
        
        # Configure column widths
        self.db_tree.column('ec_number', width=120)
        self.db_tree.column('organism', width=150)
        self.db_tree.column('substrate', width=200)
        self.db_tree.column('mutation', width=100)
        self.db_tree.column('kcat_mut', width=100)
        self.db_tree.column('Km_mut', width=100)
        self.db_tree.column('ddG', width=80)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.db_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.db_tree.xview)
        self.db_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack treeview and scrollbars
        self.db_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Bind selection event
        self.db_tree.bind('<<TreeviewSelect>>', self.on_db_tree_select)

    def update_db_info(self):
        """Update database information display"""
        if hasattr(self, 'df'):
            info = f"Entries: {len(self.df):,} | Enzymes: {self.df['ec_number'].nunique():,} | Organisms: {self.df['organism'].nunique():,}"
            self.db_info_var.set(info)

    def load_all_data(self):
        """Load all data into the database treeview"""
        # Clear previous data
        for item in self.db_tree.get_children():
            self.db_tree.delete(item)
        
        # Populate with all data (limit to 1000 for performance)
        display_df = self.df.head(1000) if len(self.df) > 1000 else self.df
        
        for _, row in display_df.iterrows():
            self.db_tree.insert('', 'end', values=(
                row.get('ec_number', 'N/A'),
                row.get('organism', 'N/A'),
                str(row.get('substrate_kinetics', 'N/A'))[:30] + '...' if len(str(row.get('substrate_kinetics', 'N/A'))) > 30 else row.get('substrate_kinetics', 'N/A'),
                row.get('clean_mut_wt', 'N/A'),
                f"{row.get('kcat_mut', 'N/A'):.4f}" if pd.notna(row.get('kcat_mut')) else 'N/A',
                f"{row.get('Km_mut', 'N/A'):.4f}" if pd.notna(row.get('Km_mut')) else 'N/A',
                f"{row.get('ddG', 'N/A'):.4f}" if pd.notna(row.get('ddG')) else 'N/A'
            ))
        
        self.status_var.set(f"Loaded {len(display_df)} entries into database view")
        if len(self.df) > 1000:
            messagebox.showinfo("Data Loaded", f"Displaying first 1000 of {len(self.df):,} total entries for performance")

    def quick_filter(self):
        """Apply quick filter to database view"""
        filter_window = tk.Toplevel(self.root)
        filter_window.title("Quick Filter")
        filter_window.geometry("400x300")
        
        # Filter controls would go here
        ttk.Label(filter_window, text="Quick filtering options will be implemented here", 
                 padding=20).pack(expand=True)

    def export_all_data(self):
        """Export all data to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                title="Export all data as CSV"
            )
            
            if filename:
                self.df.to_csv(filename, index=False)
                messagebox.showinfo("Export Successful", f"All data exported to {filename}")
                self.status_var.set(f"Data exported to {os.path.basename(filename)}")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")

    def on_db_tree_select(self, event):
        """Handle database treeview selection"""
        selection = self.db_tree.selection()
        if selection:
            item = selection[0]
            values = self.db_tree.item(item, 'values')
            self.display_selected_enzyme(values)

    def export_data(self):
        """Export data with options"""
        export_window = tk.Toplevel(self.root)
        export_window.title("Export Data")
        export_window.geometry("400x300")
        
        ttk.Label(export_window, text="Export options will be implemented here", 
                 padding=20).pack(expand=True)   

class ProteinAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Advanced Protein Sequence Analyzer")
        self.root.geometry("1400x900")
        self.root.configure(bg='#f5f5f5')
        
        # Initialize analyzer
        self.analyzer = ProteinAnalyzer()
        
        # Store temporary files
        self.temp_files = []
        
        # Initialize status variable
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - Enter protein sequence or PDB ID to begin analysis")
        
        # Store current analysis data
        self.current_analysis = None
        self.current_description = ""
        self.current_pdb_file = None
        self.current_sequence = ""
        
        # Create GUI
        self.create_gui()
    
    def create_gui(self):
        """Create the main GUI interface"""

        
        # Create main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create left and right panes
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Input section
        self.create_input_section(left_frame)
        
        # Analysis results section
        self.create_results_section(left_frame)
        
        # Visualization section
        self.create_visualization_section(right_frame)
        
        # Status bar - ADDED THIS
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def create_menu(self):
        """Create menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
    
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Load FASTA File", command=self.load_fasta_file)
        file_menu.add_command(label="Export Results", command=self.export_results)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
    
        # Analysis menu
        analysis_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Analysis", menu=analysis_menu)
        analysis_menu.add_command(label="Full Analysis", command=self.run_full_analysis)
        analysis_menu.add_command(label="Sequence Analysis Only", command=self.analyze_sequence_only)
    
        # Switch Application menu - NEW
        switch_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Switch App", menu=switch_menu)
        switch_menu.add_command(label="Go to Enzyme Kinetics Analyzer", 
                          command=self.switch_to_enzyme_kinetics)
    
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="User Guide", command=self.show_user_guide)
        help_menu.add_command(label="About", command=self.show_about)
    
    def create_input_section(self, parent):
        """Create input section"""
        input_frame = ttk.LabelFrame(parent, text="Protein Input", padding="10")
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Input method selection
        method_frame = ttk.Frame(input_frame)
        method_frame.pack(fill=tk.X, pady=5)
        
        self.input_method = tk.StringVar(value="sequence")
        ttk.Radiobutton(method_frame, text="Protein Sequence", variable=self.input_method, 
                       value="sequence", command=self.toggle_input_method).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(method_frame, text="PDB ID", variable=self.input_method, 
                       value="pdb", command=self.toggle_input_method).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(method_frame, text="FASTA Format", variable=self.input_method, 
                       value="fasta", command=self.toggle_input_method).pack(side=tk.LEFT, padx=10)
        
        # Sequence input
        self.seq_frame = ttk.Frame(input_frame)
        self.seq_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(self.seq_frame, text="Protein Sequence:").pack(anchor=tk.W)
        self.sequence_text = tk.Text(self.seq_frame, height=6, width=60, wrap=tk.WORD)
        seq_scrollbar = ttk.Scrollbar(self.seq_frame, orient=tk.VERTICAL, command=self.sequence_text.yview)
        self.sequence_text.configure(yscrollcommand=seq_scrollbar.set)
        self.sequence_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        seq_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # PDB input
        self.pdb_frame = ttk.Frame(input_frame)
        
        ttk.Label(self.pdb_frame, text="PDB ID:").pack(anchor=tk.W)
        pdb_input_frame = ttk.Frame(self.pdb_frame)
        pdb_input_frame.pack(fill=tk.X, pady=5)
        
        self.pdb_entry = ttk.Entry(pdb_input_frame, width=20)
        self.pdb_entry.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(pdb_input_frame, text="Fetch Structure", command=self.fetch_pdb).pack(side=tk.LEFT)
        
        # FASTA input
        self.fasta_frame = ttk.Frame(input_frame)
        
        ttk.Label(self.fasta_frame, text="FASTA Format:").pack(anchor=tk.W)
        self.fasta_text = tk.Text(self.fasta_frame, height=6, width=60, wrap=tk.WORD)
        fasta_scrollbar = ttk.Scrollbar(self.fasta_frame, orient=tk.VERTICAL, command=self.fasta_text.yview)
        self.fasta_text.configure(yscrollcommand=fasta_scrollbar.set)
        self.fasta_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        fasta_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons
        # In the create_input_section method of ProteinAnalyzerGUI class, modify the button_frame section:

        # Buttons
        button_frame = ttk.Frame(input_frame)
        button_frame.pack(fill=tk.X, pady=10)

        ttk.Button(button_frame, text="Clear All", command=self.clear_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Analyze Protein", command=self.run_full_analysis).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Generate Report", command=self.generate_report).pack(side=tk.LEFT, padx=5)
        # ADD THIS NEW BUTTON
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)

    # Add this new method to the ProteinAnalyzerGUI class:
    def export_to_excel(self):
        """Export analysis results to Excel"""
        if not hasattr(self, 'current_analysis'):
            messagebox.showwarning("No Data", "No analysis results to export. Please analyze a protein first.")
            return
    
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Export Analysis to Excel"
            )
        
            if filename:
                success = self.analyzer.export_to_excel(
                    self.current_analysis,
                    self.current_description if hasattr(self, 'current_description') else "",
                    filename
                )
            
                if success:
                    messagebox.showinfo("Export Successful", f"Analysis results exported to:\n{filename}")
                    self.status_var.set(f"Exported to Excel: {os.path.basename(filename)}")
                else:
                    messagebox.showerror("Export Error", "Failed to export results to Excel")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel: {str(e)}")

    def create_results_section(self, parent):
        """Create results display section"""
        results_frame = ttk.LabelFrame(parent, text="Analysis Results", padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create notebook for organized results
        self.results_notebook = ttk.Notebook(results_frame)
        self.results_notebook.pack(fill=tk.BOTH, expand=True)
        
        # Summary tab
        self.summary_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(self.summary_tab, text="Summary")
        
        self.summary_text = scrolledtext.ScrolledText(self.summary_tab, wrap=tk.WORD, width=60, height=15)
        self.summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Detailed Analysis tab
        self.details_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(self.details_tab, text="Detailed Analysis")
        
        self.details_text = scrolledtext.ScrolledText(self.details_tab, wrap=tk.WORD, width=60, height=15)
        self.details_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
    
    def create_visualization_section(self, parent):
        """Create visualization section"""
        viz_frame = ttk.LabelFrame(parent, text="Visualizations", padding="10")
        viz_frame.pack(fill=tk.BOTH, expand=True)
        
        # Visualization controls
        control_frame = ttk.Frame(viz_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(control_frame, text="Plot Type:").pack(side=tk.LEFT, padx=5)
        self.plot_type = tk.StringVar(value="composition")
        plot_combo = ttk.Combobox(control_frame, textvariable=self.plot_type,
                                 values=["Amino Acid Composition", 
                                         "Hydrophobicity Profile","Ramachandran Plot"],
                                 state="readonly")
        plot_combo.pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Generate Plot", command=self.generate_plot).pack(side=tk.LEFT, padx=10)
        
        # Plot frame
        self.plot_frame = ttk.Frame(viz_frame)
        self.plot_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # 3D structure frame
        struct_viz_frame = ttk.LabelFrame(viz_frame, text="3D Structure Viewer", padding="10")
        struct_viz_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        self.struct_viz_text = tk.Text(struct_viz_frame, height=8, wrap=tk.WORD)
        struct_viz_scrollbar = ttk.Scrollbar(struct_viz_frame, orient=tk.VERTICAL, command=self.struct_viz_text.yview)
        self.struct_viz_text.configure(yscrollcommand=struct_viz_scrollbar.set)
        self.struct_viz_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        struct_viz_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add structure viewer buttons
        struct_button_frame = ttk.Frame(struct_viz_frame)
        struct_button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(struct_button_frame, text="View 3D Structure", command=self.view_3d_structure).pack(side=tk.LEFT, padx=5)
    
    def toggle_input_method(self):
        """Toggle between input methods"""
        method = self.input_method.get()
        
        # Hide all frames
        self.seq_frame.pack_forget()
        self.pdb_frame.pack_forget()
        self.fasta_frame.pack_forget()
        
        # Show selected frame
        if method == "sequence":
            self.seq_frame.pack(fill=tk.X, pady=5)
        elif method == "pdb":
            self.pdb_frame.pack(fill=tk.X, pady=5)
        elif method == "fasta":
            self.fasta_frame.pack(fill=tk.X, pady=5)
    
    def get_sequence(self):
        """Get protein sequence from current input method"""
        method = self.input_method.get()
        sequence = ""
        description = ""
        
        try:
            if method == "sequence":
                sequence = self.sequence_text.get(1.0, tk.END).strip().upper()
                # Remove any non-amino acid characters
                sequence = re.sub(r'[^A-Z]', '', sequence)
                description = "User-provided sequence"
                self.current_sequence = sequence
                
            elif method == "pdb":
                pdb_id = self.pdb_entry.get().strip().upper()
                if not pdb_id:
                    raise ValueError("Please enter a PDB ID")
                
                self.status_var.set(f"Fetching PDB structure {pdb_id}...")
                self.root.update()
                
                sequence, pdb_file = self.analyzer.fetch_pdb_sequence(pdb_id)
                if sequence:
                    description = f"PDB ID: {pdb_id}"
                    # Store PDB file for later use
                    self.current_pdb_file = pdb_file
                    self.current_sequence = sequence
                    if pdb_file and pdb_file not in self.temp_files:
                        self.temp_files.append(pdb_file)
                else:
                    raise ValueError(f"Could not fetch sequence for PDB ID: {pdb_id}")
                    
            elif method == "fasta":
                fasta_text = self.fasta_text.get(1.0, tk.END).strip()
                if not fasta_text:
                    raise ValueError("Please enter FASTA format text")
                
                description, sequence = self.analyzer.parse_fasta(fasta_text)
                self.current_sequence = sequence
                if not sequence:
                    raise ValueError("Invalid FASTA format")
            
            # Validate sequence
            if not sequence:
                raise ValueError("No sequence found")
                
            is_valid, invalid_chars = self.analyzer.validate_sequence(sequence)
            if not is_valid:
                raise ValueError(f"Invalid amino acids in sequence: {invalid_chars}")
            
            return sequence, description
            
        except Exception as e:
            messagebox.showerror("Input Error", str(e))
            return None, None
    
    def run_full_analysis(self):
        """Run complete protein analysis"""
        try:
            sequence, description = self.get_sequence()
            if not sequence:
                return
            
            self.status_var.set("Analyzing protein sequence...")
            self.root.update()
            
            # Perform sequence analysis
            analysis = self.analyzer.analyze_sequence(sequence)
            
            if 'error' in analysis:
                messagebox.showerror("Analysis Error", analysis['error'])
                return
            
            # Display results
            self.display_analysis_results(analysis, description)
            
            # If we have a PDB file, perform structure analysis
            if hasattr(self, 'current_pdb_file') and self.current_pdb_file:
                self.status_var.set("Analyzing protein structure...")
                self.root.update()
                
                structure_analysis = self.analyzer.analyze_pdb_structure(self.current_pdb_file)
                if structure_analysis:
                    self.display_structure_analysis(structure_analysis)
            
            self.status_var.set("Analysis complete!")
            
        except Exception as e:
            messagebox.showerror("Analysis Error", f"Failed to analyze protein: {str(e)}")
            self.status_var.set("Analysis failed")
    
    def analyze_sequence_only(self):
        """Analyze sequence only"""
        try:
            sequence, description = self.get_sequence()
            if not sequence:
                return
            
            self.status_var.set("Analyzing protein sequence...")
            self.root.update()
            
            analysis = self.analyzer.analyze_sequence(sequence)
            
            if 'error' in analysis:
                messagebox.showerror("Analysis Error", analysis['error'])
                return
            
            self.display_analysis_results(analysis, description)
            self.status_var.set("Sequence analysis complete!")
            
        except Exception as e:
            messagebox.showerror("Analysis Error", f"Failed to analyze sequence: {str(e)}")
    
    def analyze_structure_only(self):
        """Analyze structure only (requires PDB input)"""
        if self.input_method.get() != "pdb":
            messagebox.showwarning("Input Required", "Structure analysis requires PDB ID input")
            return
        
        try:
            pdb_id = self.pdb_entry.get().strip().upper()
            if not pdb_id:
                messagebox.showwarning("Input Required", "Please enter a PDB ID")
                return
            
            self.status_var.set(f"Fetching and analyzing PDB structure {pdb_id}...")
            self.root.update()
            
            # Fetch structure
            structure, pdb_file = self.analyzer.fetch_pdb_structure(pdb_id)
            if not structure:
                raise ValueError(f"Could not fetch PDB structure: {pdb_id}")
            
            # Store PDB file
            self.current_pdb_file = pdb_file
            if pdb_file and pdb_file not in self.temp_files:
                self.temp_files.append(pdb_file)
            
            # Analyze structure
            structure_analysis = self.analyzer.analyze_pdb_structure(pdb_file)
            if structure_analysis:
                self.display_structure_analysis(structure_analysis)
                self.status_var.set("Structure analysis complete!")
            else:
                raise ValueError("Structure analysis failed")
                
        except Exception as e:
            messagebox.showerror("Structure Analysis Error", f"Failed to analyze structure: {str(e)}")
    
    def display_analysis_results(self, analysis, description):
        """Display analysis results in the GUI"""
        # Summary tab
        summary_report = self.analyzer.format_analysis_report(analysis, description)
        self.summary_text.config(state=tk.NORMAL)
        self.summary_text.delete(1.0, tk.END)
        self.summary_text.insert(1.0, summary_report)
        self.summary_text.config(state=tk.DISABLED)
        
        # Detailed analysis tab
        details = self.format_detailed_analysis(analysis)
        self.details_text.config(state=tk.NORMAL)
        self.details_text.delete(1.0, tk.END)
        self.details_text.insert(1.0, details)
        self.details_text.config(state=tk.DISABLED)
        
        # Store current analysis for plotting
        self.current_analysis = analysis
        self.current_description = description
    
    def display_structure_analysis(self, structure_analysis):
        """Display structure analysis results"""
        structure_text = "PROTEIN STRUCTURE ANALYSIS\n"
        structure_text += "=" * 50 + "\n\n"
        
        structure_text += f"Number of models: {structure_analysis.get('num_models', 'N/A')}\n"
        structure_text += f"Number of chains: {structure_analysis.get('num_chains', 'N/A')}\n"
        structure_text += f"Number of residues: {structure_analysis.get('num_residues', 'N/A')}\n"
        structure_text += f"Number of atoms: {structure_analysis.get('num_atoms', 'N/A')}\n\n"
        
        if 'center_of_mass' in structure_analysis:
            com = structure_analysis['center_of_mass']
            structure_text += f"Center of mass: ({com[0]:.2f}, {com[1]:.2f}, {com[2]:.2f}) Å\n"
        
        if 'dimensions' in structure_analysis:
            dims = structure_analysis['dimensions']
            structure_text += f"Dimensions: {dims[0]:.2f} × {dims[1]:.2f} × {dims[2]:.2f} Å\n"
        
        self.structure_text.config(state=tk.NORMAL)
        self.structure_text.delete(1.0, tk.END)
        self.structure_text.insert(1.0, structure_text)
        self.structure_text.config(state=tk.DISABLED)
    
    def format_detailed_analysis(self, analysis):
        """Format detailed analysis results"""
        details = "DETAILED PROTEIN ANALYSIS\n"
        details += "=" * 50 + "\n\n"
        
        # Amino acid composition in detail
        details += "AMINO ACID COMPOSITION (Detailed):\n"
        comp = analysis['amino_acid_composition']
        sorted_comp = sorted(comp.items(), key=lambda x: x[1], reverse=True)
        
        # Group by properties
        hydrophobic = 'AVILMFYW'  # Hydrophobic
        polar = 'STNQ'  # Polar uncharged
        positive = 'KRH'  # Positively charged
        negative = 'DE'  # Negatively charged
        special = 'CGP'  # Special cases
        
        categories = {
            'Hydrophobic': [],
            'Polar Uncharged': [],
            'Positively Charged': [],
            'Negatively Charged': [],
            'Special': []
        }
        
        for aa, perc in sorted_comp:
            if aa in hydrophobic:
                categories['Hydrophobic'].append((aa, perc))
            elif aa in polar:
                categories['Polar Uncharged'].append((aa, perc))
            elif aa in positive:
                categories['Positively Charged'].append((aa, perc))
            elif aa in negative:
                categories['Negatively Charged'].append((aa, perc))
            elif aa in special:
                categories['Special'].append((aa, perc))
        
        for category, aas in categories.items():
            if aas:
                details += f"\n{category}:\n"
                for aa, perc in aas:
                    details += f"  {aa}: {perc:.2f}%\n"
        
        # Charge properties
        details += f"\nCHARGE PROPERTIES:\n"
        details += f"  Net charge at pH 7.0: {analysis['charge_at_ph7']:.3f}\n"
        details += f"  Isoelectric point (pI): {analysis['isoelectric_point']:.2f}\n"
        
        # Stability and function
        details += f"\nSTABILITY AND FUNCTION:\n"
        details += f"  Instability index: {analysis['instability_index']:.2f}\n"
        half_life_hours = analysis['half_life'] / 3600
        details += f"  Predicted half-life: {half_life_hours:.1f} hours\n"
        details += f"  Extinction coefficient: {analysis['extinction_coefficient']:.0f} M⁻¹cm⁻¹\n"
        
        # Hydrophobicity
        details += f"\nHYDROPHOBICITY:\n"
        details += f"  Average hydrophobicity: {analysis['hydrophobicity']:.3f}\n"
        
        # Secondary structure prediction
        ss = analysis['secondary_structure']
        details += f"\nSECONDARY STRUCTURE PREDICTION:\n"
        details += f"  α-helix: {ss['helix']:.1f}%\n"
        details += f"  β-sheet: {ss['sheet']:.1f}%\n"
        details += f"  Random coil: {ss['coil']:.1f}%\n"
        
        return details
    
    def generate_plot(self):
        """Generate selected plot"""
        if not hasattr(self, 'current_analysis'):
            messagebox.showwarning("No Data", "Please analyze a protein first")
            return
        
        # Clear previous plot
        for widget in self.plot_frame.winfo_children():
            widget.destroy()
        
        plot_type = self.plot_type.get()
        
        try:
            fig = None
            
            if plot_type == "Amino Acid Composition":
                fig = self.plot_amino_acid_composition()
            elif plot_type == "Hydrophobicity Profile":
                fig = self.plot_hydrophobicity_profile()
            elif plot_type == "Ramachandran Plot":
                fig = self.plot_ramachandran()

            
            if fig:
                # Embed plot in tkinter
                canvas = FigureCanvasTkAgg(fig, self.plot_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
                # Add toolbar
                from matplotlib.backends.backend_tkagg import NavigationToolbar2Tk
                toolbar = NavigationToolbar2Tk(canvas, self.plot_frame)
                toolbar.update()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
                self.status_var.set(f"Generated {plot_type} plot")
            
        except Exception as e:
            messagebox.showerror("Plot Error", f"Failed to generate plot: {str(e)}")
    
    def plot_amino_acid_composition(self):
        """Plot amino acid composition"""
        fig, ax = plt.subplots(figsize=(10, 6))
        
        comp = self.current_analysis['amino_acid_composition']
        amino_acids = list(comp.keys())
        percentages = list(comp.values())
        
        # Sort by percentage
        sorted_indices = np.argsort(percentages)[::-1]
        amino_acids = [amino_acids[i] for i in sorted_indices]
        percentages = [percentages[i] for i in sorted_indices]
        
        bars = ax.bar(amino_acids, percentages, color='skyblue', edgecolor='black')
        
        # Add value labels on bars
        for bar, percentage in zip(bars, percentages):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{percentage:.1f}%', ha='center', va='bottom')
        
        ax.set_xlabel('Amino Acid')
        ax.set_ylabel('Percentage (%)')
        ax.set_title('Amino Acid Composition')
        ax.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        return fig
    
    def plot_hydrophobicity_profile(self):
        """Plot hydrophobicity profile along the sequence"""
        fig, ax = plt.subplots(figsize=(12, 4))
        
        # Get sequence from analysis
        if hasattr(self, 'current_sequence'):
            sequence = self.current_sequence
        else:
            # For demonstration, use a dummy sequence
            sequence = "ACDEFGHIKLMNPQRSTVWY" * 10  # Repeat standard amino acids
        
        # Calculate hydrophobicity for each position
        hydrophobicity_values = []
        for aa in sequence[:100]:  # Limit to first 100 residues for clarity
            if aa in self.analyzer.hydrophobicity:
                hydrophobicity_values.append(self.analyzer.hydrophobicity[aa])
            else:
                hydrophobicity_values.append(0)
        
        positions = range(1, len(hydrophobicity_values) + 1)
        
        ax.plot(positions, hydrophobicity_values, linewidth=2, color='blue')
        ax.fill_between(positions, hydrophobicity_values, alpha=0.3, color='lightblue')
        
        ax.axhline(y=0, color='red', linestyle='--', alpha=0.7, label='Hydrophilic/Hydrophobic boundary')
        
        ax.set_xlabel('Residue Position')
        ax.set_ylabel('Hydrophobicity (Kyte-Doolittle)')
        ax.set_title('Hydrophobicity Profile')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        return fig
    
    def plot_ramachandran(self):
        """Plot Ramachandran plot"""
        fig, ax = plt.subplots(figsize=(5, 5))
        
        if 'ramachandran_data' in self.current_analysis:
            phi = self.current_analysis['ramachandran_data']['phi']
            psi = self.current_analysis['ramachandran_data']['psi']
            residues = self.current_analysis['ramachandran_data']['residues']
            
            # Color by residue type
            colors = []
            for res in residues:
                if res in 'G':  # Glycine
                    colors.append('red')
                elif res in 'P':  # Proline
                    colors.append('green')
                else:
                    colors.append('blue')
            
            scatter = ax.scatter(phi, psi, c=colors, alpha=0.6, s=30)
            
            # Add region boundaries
            ax.axvline(0, color='gray', linestyle='--', alpha=0.5)
            ax.axhline(0, color='gray', linestyle='--', alpha=0.5)
            
            # Add region labels
            ax.text(-150, 150, 'β-sheet', fontsize=12, ha='center', color='darkblue')
            ax.text(-100, -50, 'α-helix', fontsize=12, ha='center', color='darkred')
            ax.text(100, 50, 'Left-handed\nhelix', fontsize=10, ha='center', color='darkgreen')
            
            ax.set_xlabel('Phi Angle (degrees)')
            ax.set_ylabel('Psi Angle (degrees)')
            ax.set_title('Ramachandran Plot')
            ax.set_xlim(-180, 180)
            ax.set_ylim(-180, 180)
            ax.grid(True, alpha=0.3)
            
            # Add legend
            from matplotlib.lines import Line2D
            legend_elements = [
                Line2D([0], [0], marker='o', color='w', markerfacecolor='blue', markersize=8, label='Standard'),
                Line2D([0], [0], marker='o', color='w', markerfacecolor='red', markersize=8, label='Glycine'),
                Line2D([0], [0], marker='o', color='w', markerfacecolor='green', markersize=8, label='Proline')
            ]
            ax.legend(handles=legend_elements, loc='upper right')
        
        else:
            ax.text(0.5, 0.5, 'Ramachandran data not available', 
                   ha='center', va='center', transform=ax.transAxes, fontsize=12)
        
        plt.tight_layout()
        return fig
    
    
    def view_3d_structure(self):
        """View 3D structure in browser"""
        if not hasattr(self, 'current_pdb_file') or not self.current_pdb_file:
            messagebox.showwarning("No Structure", "No PDB structure available. Please fetch a PDB structure first.")
            return
        
        try:
            pdb_id = "PREDICTED"
            if hasattr(self, 'current_description') and 'PDB ID:' in self.current_description:
                pdb_id = self.current_description.split('PDB ID:')[-1].strip()
            
            html_file = self.analyzer.visualize_structure(self.current_pdb_file, pdb_id)
            if html_file:
                # Open in default browser
                webbrowser.open('file://' + os.path.abspath(html_file))
                self.temp_files.append(html_file)
                self.status_var.set("Opened 3D structure in browser")
            else:
                raise ValueError("Could not generate 3D visualization")
                
        except Exception as e:
            messagebox.showerror("3D View Error", f"Failed to open 3D structure: {str(e)}")
    
    def load_fasta_file(self):
        """Load FASTA file from disk"""
        filename = filedialog.askopenfilename(
            title="Select FASTA File",
            filetypes=[("FASTA files", "*.fasta;*.fa"), ("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, 'r') as file:
                    fasta_content = file.read()
                
                self.input_method.set("fasta")
                self.toggle_input_method()
                self.fasta_text.delete(1.0, tk.END)
                self.fasta_text.insert(1.0, fasta_content)
                
                self.status_var.set(f"Loaded FASTA file: {os.path.basename(filename)}")
                
            except Exception as e:
                messagebox.showerror("File Error", f"Failed to load FASTA file: {str(e)}")
    
    def export_results(self):
        """Export analysis results"""
        if not hasattr(self, 'current_analysis'):
            messagebox.showwarning("No Data", "No analysis results to export")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("Text files", "*.txt"), ("All files", "*.*")],
            title="Export Analysis Results"
        )
        
        if filename:
            try:
                if filename.endswith('.xlsx'):
                    success = self.analyzer.export_to_excel(
                        self.current_analysis, 
                        self.current_description if hasattr(self, 'current_description') else "",
                        filename
                    )
                    if success:
                        messagebox.showinfo("Export Successful", f"Results exported to {filename}")
                        self.status_var.set(f"Exported results to {os.path.basename(filename)}")
                    else:
                        raise ValueError("Excel export failed")
                else:
                    # Export as text
                    report = self.analyzer.format_analysis_report(
                        self.current_analysis,
                        self.current_description if hasattr(self, 'current_description') else ""
                    )
                    with open(filename, 'w') as file:
                        file.write(report)
                    messagebox.showinfo("Export Successful", f"Results exported to {filename}")
                    self.status_var.set(f"Exported results to {os.path.basename(filename)}")
                    
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export results: {str(e)}")
    
    def generate_report(self):
        """Generate comprehensive analysis report"""
        if not hasattr(self, 'current_analysis'):
            messagebox.showwarning("No Data", "No analysis results to report")
            return
        
        # This could generate a more comprehensive PDF report
        # For now, just show a message
        messagebox.showinfo("Report Generation", 
                          "Comprehensive report generation feature will be implemented in future versions.\n"
                          "Use 'Export Results' to save current analysis.")
    
    def clear_all(self):
        """Clear all inputs and results"""
        self.sequence_text.delete(1.0, tk.END)
        self.pdb_entry.delete(0, tk.END)
        self.fasta_text.delete(1.0, tk.END)
        
        self.summary_text.config(state=tk.NORMAL)
        self.summary_text.delete(1.0, tk.END)
        self.summary_text.config(state=tk.DISABLED)
        
        self.details_text.config(state=tk.NORMAL)
        self.details_text.delete(1.0, tk.END)
        self.details_text.config(state=tk.DISABLED)
        
        self.structure_text.config(state=tk.NORMAL)
        self.structure_text.delete(1.0, tk.END)
        self.structure_text.config(state=tk.DISABLED)
        
        self.struct_viz_text.config(state=tk.NORMAL)
        self.struct_viz_text.delete(1.0, tk.END)
        self.struct_viz_text.config(state=tk.DISABLED)
        
        # Clear plot frame
        for widget in self.plot_frame.winfo_children():
            widget.destroy()
        
        # Clear temporary attributes
        if hasattr(self, 'current_analysis'):
            del self.current_analysis
        if hasattr(self, 'current_description'):
            del self.current_description
        if hasattr(self, 'current_pdb_file'):
            del self.current_pdb_file
        if hasattr(self, 'current_sequence'):
            del self.current_sequence
        
        self.status_var.set("All inputs and results cleared")
    
    def fetch_pdb(self):
        """Fetch PDB structure"""
        pdb_id = self.pdb_entry.get().strip().upper()
        if not pdb_id:
            messagebox.showwarning("Input Required", "Please enter a PDB ID")
            return
        
        try:
            self.status_var.set(f"Fetching PDB structure {pdb_id}...")
            self.root.update()
            
            structure, pdb_file = self.analyzer.fetch_pdb_structure(pdb_id)
            if structure:
                self.current_pdb_file = pdb_file
                if pdb_file and pdb_file not in self.temp_files:
                    self.temp_files.append(pdb_file)
                
                # Get sequence from structure
                sequence, _ = self.analyzer.fetch_pdb_sequence(pdb_id)
                if sequence:
                    self.current_sequence = sequence
                    # Update sequence display
                    self.sequence_text.delete(1.0, tk.END)
                    self.sequence_text.insert(1.0, sequence)
                
                messagebox.showinfo("Success", f"PDB structure {pdb_id} fetched successfully!")
                self.status_var.set(f"PDB structure {pdb_id} fetched and ready for analysis")
            else:
                raise ValueError(f"Could not fetch PDB structure: {pdb_id}")
                
        except Exception as e:
            messagebox.showerror("Fetch Error", f"Failed to fetch PDB structure: {str(e)}")
            self.status_var.set("PDB fetch failed")
    
    def show_user_guide(self):
        """Show user guide"""
        guide_text = """
PROTEIN SEQUENCE ANALYZER - USER GUIDE

1. INPUT METHODS:
   - Protein Sequence: Enter raw amino acid sequence (single letter code)
   - PDB ID: Enter 4-character PDB identifier to fetch structure
   - FASTA Format: Enter sequence in FASTA format (starts with '>')

2. ANALYSIS OPTIONS:
   - Full Analysis: Complete sequence and structure analysis
   - Sequence Analysis Only: Analyze sequence properties only
   - Structure Analysis Only: Analyze 3D structure (requires PDB ID)

3. RESULTS:
   - Summary: Basic protein properties and statistics
   - Detailed Analysis: Comprehensive analysis with categories
   - Structure Analysis: 3D structure information

4. VISUALIZATIONS:
   - Amino Acid Composition: Bar chart of amino acid percentages
   - Hydrophobicity Profile: Hydrophobicity along sequence
   - Secondary Structure: Pie chart of predicted structure
   - Ramachandran Plot: Phi/Psi angles distribution
   - Charge Distribution: Net charge vs pH

5. 3D STRUCTURE:
   - View 3D Structure: Open structure in browser (requires PDB)
   - Predict Structure: Generate predicted structure from sequence

6. EXPORT:
   - Export results to Excel or text format

TIPS:
- For best results with PDB structures, ensure stable internet connection
- Sequence analysis works offline
- Large structures may take longer to load and analyze
- Use 'Clear All' to reset the interface
"""
        messagebox.showinfo("User Guide", guide_text)
    
    def show_about(self):
        """Show about information"""
        about_text = """
Advanced Protein Sequence Analyzer

Version: 1.0
Developed for: Protein Bioinformatics Research

Features:
- Comprehensive protein sequence analysis
- PDB structure fetching and analysis
- Multiple visualization types
- 3D structure viewing
- Property predictions
- Export capabilities

This tool is designed for researchers studying protein structure, 
function, and sequence-structure relationships.
"""
        messagebox.showinfo("About", about_text)
    
    def __del__(self):
        """Cleanup temporary files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass

class ApplicationSelector:
    def __init__(self, root):
        self.root = root
        self.root.title("Bioinformatics Tools Suite")
        self.root.geometry("600x400")
        self.root.configure(bg='#f0f0f0')
        
        # Center the window
        self.center_window(600, 400)
        
        self.create_gui()
    
    def center_window(self, width, height):
        """Center the window on screen"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_gui(self):
        """Create the application selection GUI"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="40")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Bioinformatics Tools Suite", 
                               font=('Arial', 20, 'bold'))
        title_label.pack(pady=20)
        
        # Description
        desc_label = ttk.Label(main_frame, 
                              text="Select an application to launch:",
                              font=('Arial', 12))
        desc_label.pack(pady=10)
        
        # Application buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=30, fill=tk.BOTH, expand=True)
        
        # Protein Analyzer Button
        protein_btn = tk.Button(button_frame, 
                               text="🧬 Protein Sequence Analyzer", 
                               font=('Arial', 14, 'bold'),
                               bg='#4CAF50', fg='white',
                               width=25, height=3,
                               command=self.launch_protein_analyzer)
        protein_btn.pack(pady=15)
        
        # Enzyme Kinetics Button
        enzyme_btn = tk.Button(button_frame,
                              text="⚡ Enzyme Kinetics Analyzer", 
                              font=('Arial', 14, 'bold'),
                              bg='#2196F3', fg='white',
                              width=25, height=3,
                              command=self.launch_enzyme_kinetics)
        enzyme_btn.pack(pady=15)
        
        # Info frame
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=20)
        
        info_text = """
        Protein Sequence Analyzer: Comprehensive analysis of protein sequences including 
        physicochemical properties, structure predictions, and 3D visualization.
        
        Enzyme Kinetics Analyzer: Advanced analysis of enzyme kinetics data with 
        statistical analysis, visualization, and database management.
        """
        
        info_label = ttk.Label(info_frame, text=info_text, 
                              font=('Arial', 9), justify=tk.CENTER,
                              background='#f0f0f0')
        info_label.pack()
    
    def launch_protein_analyzer(self):
        """Launch the Protein Sequence Analyzer"""
        self.root.withdraw()  # Hide selector window
        protein_window = tk.Toplevel(self.root)
        protein_window.title("Advanced Protein Sequence Analyzer")
        protein_window.geometry("1400x900")
        
        # Center the new window
        self.center_child_window(protein_window, 1400, 900)
        
        app = ProteinAnalyzerGUI(protein_window)
        
        # Handle window close
        protein_window.protocol("WM_DELETE_WINDOW", 
                               lambda: self.on_child_close(protein_window))
    
    def launch_enzyme_kinetics(self):
        """Launch the Enzyme Kinetics Analyzer"""
        self.root.withdraw()  # Hide selector window
        enzyme_window = tk.Toplevel(self.root)
        enzyme_window.title("Advanced Enzyme Kinetics Database Analyzer")
        enzyme_window.geometry("1600x1000")
        
        # Center the new window
        self.center_child_window(enzyme_window, 1600, 1000)
        
        app = EnzymeKineticsAnalyzer(enzyme_window)
        
        # Handle window close
        enzyme_window.protocol("WM_DELETE_WINDOW", 
                              lambda: self.on_child_close(enzyme_window))
    
    def center_child_window(self, window, width, height):
        """Center a child window relative to the main window"""
        # Get main window position
        main_x = self.root.winfo_x()
        main_y = self.root.winfo_y()
        main_width = self.root.winfo_width()
        main_height = self.root.winfo_height()
        
        # Calculate position to center relative to main window
        x = main_x + (main_width - width) // 2
        y = main_y + (main_height - height) // 2
        
        window.geometry(f'{width}x{height}+{x}+{y}')
    
    def on_child_close(self, child_window):
        """Handle child window closing"""
        child_window.destroy()
        self.root.deiconify()  # Show selector window again

def main():
    """Main function to launch the application selector"""
    root = tk.Tk()
    app = ApplicationSelector(root)
    root.mainloop()

if __name__ == "__main__":
    main()
